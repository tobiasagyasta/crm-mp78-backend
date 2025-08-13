from flask import Blueprint, request, jsonify, send_file
from decimal import Decimal
from flask_cors import cross_origin
from datetime import datetime, timedelta
from io import BytesIO
from app.models.outlet import Outlet
from app.models.gojek_reports import GojekReport
from app.models.shopee_reports import ShopeeReport
from app.models.grabfood_reports import GrabFoodReport
from app.models.cash_reports import CashReport
from app.models.manual_entry import ManualEntry
from app.models.shopeepay_reports import ShopeepayReport
from app.models.tiktok_reports import TiktokReport
from app.models.bank_mutations import BankMutation
from app.models.pukis import Pukis
from app.utils.transaction_matcher import TransactionMatcher
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from collections import defaultdict
from openpyxl.workbook.protection import WorkbookProtection
import os

export_bp = Blueprint('export', __name__, url_prefix="/export")

@export_bp.route('', methods=['POST', 'OPTIONS'])
@cross_origin(expose_headers=["Content-Disposition"])
def export_reports():
    try:

          # Define yellow background for grand total
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow background
        if request.method == 'OPTIONS':
            response = jsonify({'status': 'OK'})
            response.headers.add('Access-Control-Allow-Headers', 'Content-Type')
            response.headers.add('Access-Control-Allow-Methods', 'POST')
            # response.headers.add('Access-Control-Allow-Origin', '*')
            return response, 200

        data = request.get_json()
        outlet_code = data.get('outlet_code')
        start_date = datetime.strptime(data.get('start_date'), '%Y-%m-%d')
        end_date = datetime.strptime(data.get('end_date'), '%Y-%m-%d')

        if not all([outlet_code, start_date, end_date]):
            return jsonify({"error": "Missing required parameters"}), 400

        # Get outlet
        outlet = Outlet.query.filter_by(outlet_code=outlet_code).first()
        if not outlet:
            return jsonify({"error": "Outlet not found"}), 404

        # Initialize daily_totals dictionary
        daily_totals = {}

        # Initialize dictionary structure
        def init_daily_total():
            return {
                'Gojek_Gross': 0, 'Gojek_Net': 0,
                'Grab_Gross': 0, 'Grab_Net': 0,
                'ShopeePay_Gross': 0, 'ShopeePay_Net': 0,
                'Shopee_Gross': 0, 'Shopee_Net': 0,
                'Tiktok_Gross':0, 'Tiktok_Net': 0,
                'Cash_Income': 0, 'Cash_Expense': 0,
                'Gojek_Mutation': None, 'Gojek_Difference': 0,
                # 'Grab_Mutation': None, 
                'Grab_Difference': 0,
                'Grab_Commission': 0,
                'Shopee_Mutation': None, 'Shopee_Difference': 0,
                'ShopeePay_Mutation': None, 'ShopeePay_Difference': 0
            }
        all_dates = []
        date = start_date.date()
        end = end_date.date()
        while date <= end:
            all_dates.append(date)
            if date not in daily_totals:
                daily_totals[date] = init_daily_total()
            date += timedelta(days=1)
        
        # Add title rows
        dataset = []
        dataset.append(['Sales Report', '', '', '', '', '', '', '', ''])
        dataset.append(['Period:', f'{start_date.strftime("%Y-%m-%d")} to {end_date.strftime("%Y-%m-%d")}', '', '', '', '', '', '', ''])
        dataset.append(['Outlet:', outlet.outlet_name_gojek, '', '', '', '', '', '', ''])
        dataset.append([]) # Empty row for spacing
        # Update the initial headers to include ShopeePay, mutation data, and Minusan (Mutasi), but remove Grab Mutation and Grab Difference
        dataset.append([
            'Date',
            'Gojek Net', 'Gojek Mutation', 'Gojek Difference',
            'Grab Net',
            'Grab Net (after commission)',
            'Shopee Net', 'Shopee Mutation', 'Shopee Difference',
            'ShopeePay Net','ShopeePay Mutation', 'ShopeePay Difference','Tiktok Net','Cash Income (Admin)', 'Cash Expense (Admin)', 'Sisa Cash (Admin)','Minusan (Mutasi)'
        ])

        # Query reports with inclusive end date
        end_date_inclusive = datetime(end_date.year, end_date.month, end_date.day, 23, 59, 59)
        
        gojek_reports = GojekReport.query.filter(
            GojekReport.outlet_code == outlet_code,
            GojekReport.transaction_date >= start_date,
            GojekReport.transaction_date <= end_date_inclusive
        ).all()

        grab_reports = GrabFoodReport.query.filter(
            GrabFoodReport.outlet_code == outlet_code,
            GrabFoodReport.tanggal_dibuat >= start_date,
            GrabFoodReport.tanggal_dibuat <= end_date_inclusive
        ).all()

        shopee_reports = ShopeeReport.query.filter(
            ShopeeReport.outlet_code == outlet_code,
            ShopeeReport.order_create_time >= start_date,
            ShopeeReport.order_create_time <= end_date_inclusive
        ).all()

        shopeepay_reports = ShopeepayReport.query.filter(
            ShopeepayReport.outlet_code == outlet_code,
            ShopeepayReport.create_time >= start_date,
            ShopeepayReport.create_time <= end_date_inclusive
        ).all()
        tiktok_reports = TiktokReport.query.filter(
            TiktokReport.outlet_code == outlet_code,
            TiktokReport.order_time >= start_date,
            TiktokReport.order_time <= end_date_inclusive
        ).all()

        # Query cash reports with separate income and expense queries
        cash_income_reports = CashReport.query.filter(
            CashReport.outlet_code == outlet_code,
            CashReport.type == 'income',
            CashReport.tanggal >= start_date,
            CashReport.tanggal <= end_date_inclusive
        ).all()
        
        cash_expense_reports = CashReport.query.filter(
            CashReport.outlet_code == outlet_code,
            CashReport.type == 'expense',
            CashReport.tanggal >= start_date,
            CashReport.tanggal <= end_date_inclusive
        ).all()

        if outlet.brand == "Pukis & Martabak Kota Baru":
            pukis_reports = Pukis.query.filter(
                Pukis.outlet_code == outlet_code,
                Pukis.tanggal >= start_date,
                Pukis.tanggal <= end_date_inclusive
            ).order_by(Pukis.tanggal.asc()).all()

        # Aggregate all data by date
        for report in gojek_reports:
            date = report.transaction_date
            if date not in daily_totals:
                daily_totals[date] = init_daily_total()
            daily_totals[date]['Gojek_Net'] += float(report.nett_amount or 0)
            daily_totals[date]['Gojek_Gross'] += float(report.amount or 0)

        # Initialize variables for GrabFood and GrabOVO
        grabfood_gross_total = 0
        grabovo_gross_total = 0
        grabfood_net_total = 0
        grabovo_net_total = 0
        tiktok_gross_total = 0

        for report in grab_reports:
            date = report.tanggal_dibuat.date()
            if date not in daily_totals:
                daily_totals[date] = init_daily_total()
            daily_totals[date]['Grab_Net'] += float(report.total or 0)
            daily_totals[date]['Grab_Gross'] += float(report.amount or 0)
            # Track separate totals for GrabOVO and GrabFood
            if hasattr(report, 'jenis'):
                if report.jenis == 'OVO':
                    grabovo_gross_total += float(report.amount or 0)
                    grabovo_net_total += float(report.total or 0)
                elif report.jenis == 'GrabFood':
                    grabfood_gross_total += float(report.amount or 0)
                    grabfood_net_total += float(report.total or 0)

        # Calculate Grab_Commission for each date after all Grab_Net values are summed
        for date in daily_totals:
            if outlet.brand not in ["Pukis & Martabak Kota Baru"]:
                daily_totals[date]['Grab_Commission'] = daily_totals[date]['Grab_Net'] * 1/74
            else:
                daily_totals[date]['Grab_Commission'] = 0

        for report in shopee_reports:
            if report.order_status != "Cancelled":
                date = report.order_create_time.date()
                if date not in daily_totals:
                    daily_totals[date] = init_daily_total()
                daily_totals[date]['Shopee_Net'] += float(report.net_income or 0)
                daily_totals[date]['Shopee_Gross'] += float(report.order_amount or 0)
        for report in shopeepay_reports:
            if report.transaction_type != "Withdrawal":
                date = report.create_time.date()
                if date not in daily_totals:
                    daily_totals[date] = init_daily_total()
                daily_totals[date]['ShopeePay_Net'] += float(report.settlement_amount or 0)
                daily_totals[date]['ShopeePay_Gross'] += float(report.transaction_amount or 0)
        for report in tiktok_reports:
            date = report.order_time.date()
            if date not in daily_totals:
                daily_totals[date] = init_daily_total()
            daily_totals[date]['Tiktok_Net'] += float(report.net_amount or 0)
            daily_totals[date]['Tiktok_Gross'] += float(report.gross_amount or 0)
        # Handle cash reports separately for income and expense
        for report in cash_income_reports:
            date = report.tanggal.date()
            if date not in daily_totals:
                daily_totals[date] = init_daily_total()
            daily_totals[date]['Cash_Income'] += float(report.total or 0)

        for report in cash_expense_reports:
            date = report.tanggal.date()
            if date not in daily_totals:
                daily_totals[date] = init_daily_total()
            daily_totals[date]['Cash_Expense'] += float(report.total or 0)

        # Add mutation matching logic for each platform
        platforms = ['gojek', 'grab', 'shopee','shopeepay']
        # Prepare a dict to collect mutations grouped by rekening_number and date
        mutations_by_platform = {}
        for platform in platforms:
            try:
                matcher = TransactionMatcher(platform)
                mutations = matcher.get_mutations_query(start_date.date(), end_date.date()).all()
                # Group mutations by rekening_number and date for this platform
                from collections import defaultdict
                if platform not in mutations_by_platform:
                    mutations_by_platform[platform] = defaultdict(lambda: defaultdict(list))
                for m in mutations:
                    date = m.tanggal if hasattr(m, 'tanggal') else None
                    rekening = m.rekening_number if hasattr(m, 'rekening_number') else None
                    if date and rekening:
                        mutations_by_platform[platform][rekening][date].append(m)
              
                for date, totals in daily_totals.items():
                    # Handle platform name capitalization properly
                    if platform == 'shopeepay':
                        platform_net_key = 'ShopeePay_Net'
                        platform_mutation_key = 'ShopeePay_Mutation'
                        platform_difference_key = 'ShopeePay_Difference'
                    else:
                        platform_net_key = f'{platform.capitalize()}_Net'
                        platform_mutation_key = f'{platform.capitalize()}_Mutation'
                        platform_difference_key = f'{platform.capitalize()}_Difference'
                    
                    if platform_net_key not in totals or totals[platform_net_key] == 0:
                        continue

                    class MockDailyTotal:
                        def __init__(self, outlet_id, date, total_net):
                            self.outlet_id = outlet_id
                            self.date = date
                            self.total_net = total_net

                    mock_total = MockDailyTotal(outlet_code, date, totals[platform_net_key])
                    

                    # Default matching for gojek/grab
                    platform_data, mutation_data = matcher.match_transactions(mock_total, mutations)
                    if mutation_data:
                        mutation_amount = float(mutation_data.get('transaction_amount', 0))
                        totals[platform_mutation_key] = mutation_amount
                        totals[platform_difference_key] = float(mutation_amount) - float(totals[platform_net_key])
                    else:
                        totals[platform_mutation_key] = None
                        totals[platform_difference_key] = None

            except Exception as e:
                print(f"Warning: Mutation matching failed for {platform}: {str(e)}")
                continue

        # Add aggregated data to dataset, using all_dates to ensure all dates are present
        from app.utils.pkb_mutation import get_minus_manual_entries
        minusan_entries = get_minus_manual_entries(outlet_code, start_date.date(), end_date_inclusive.date())
        # Build a lookup: date -> sum of minusan amounts for that date
        minusan_by_date = {}
        for entry in minusan_entries:
            d = getattr(entry, 'minus_date', None)
            if d:
                minusan_by_date.setdefault(d, 0)
                minusan_by_date[d] += float(entry.amount or 0) * -1

    

        for date in all_dates:
            totals = daily_totals[date]
            minusan_total = minusan_by_date.get(date, 0)
            cash_income = totals['Cash_Income']
            cash_expense = totals['Cash_Expense']
            daily_commission = (totals['Grab_Net'] - (totals['Grab_Net'] * 1/74)) if outlet.brand not in ["Pukis & Martabak Kota Baru"] else 0
            totals['Grab_Commission'] = daily_commission
            dataset.append([
                date,
                totals['Gojek_Net'],
                totals['Gojek_Mutation'],
                totals['Gojek_Difference'],
                totals['Grab_Net'],
                totals['Grab_Commission'],
                totals['Shopee_Net'],
                totals['Shopee_Mutation'],
                totals['Shopee_Difference'],
                totals['ShopeePay_Net'],
                totals['ShopeePay_Mutation'],
                totals['ShopeePay_Difference'],
                totals['Tiktok_Net'],   
                cash_income,
                cash_expense,
                cash_income - cash_expense,
                minusan_total,
            ])
        
        cash_income = sum(day['Cash_Income'] for day in daily_totals.values())
        cash_expense = sum(day['Cash_Expense'] for day in daily_totals.values())        # Update grand totals to include ShopeePay and mutation data
        grand_totals = {
            'Gojek_Gross': sum(day['Gojek_Gross'] for day in daily_totals.values()),
            'Gojek_Net': sum(day['Gojek_Net'] for day in daily_totals.values()),
            'Gojek_Mutation': sum(day['Gojek_Mutation'] for day in daily_totals.values() if day['Gojek_Mutation'] is not None),
            'Gojek_Difference': sum(day['Gojek_Difference'] for day in daily_totals.values() if day['Gojek_Difference'] is not None),
            'Grab_Gross': sum(day['Grab_Gross'] for day in daily_totals.values()),
            'Grab_Net': sum(day['Grab_Net'] for day in daily_totals.values()),
            # 'Grab_Mutation': sum(day['Grab_Mutation'] for day in daily_totals.values() if day['Grab_Mutation'] is not None),
            'Grab_Commission': sum(day['Grab_Commission'] for day in daily_totals.values() if day['Grab_Commission'] is not None),
            # 'Grab_Difference': sum(day['Grab_Difference'] for day in daily_totals.values() if day['Grab_Difference'] is not None),
            'Shopee_Gross': sum(day['Shopee_Gross'] for day in daily_totals.values()),
            'Shopee_Net': sum(day['Shopee_Net'] for day in daily_totals.values()),
            'Shopee_Mutation': sum(day['Shopee_Mutation'] for day in daily_totals.values() if day['Shopee_Mutation'] is not None),
            'Shopee_Difference': sum(day['Shopee_Difference'] for day in daily_totals.values() if day['Shopee_Difference'] is not None),
            'ShopeePay_Gross': sum(day['ShopeePay_Gross'] for day in daily_totals.values()),
            'ShopeePay_Net': sum(day['ShopeePay_Net'] for day in daily_totals.values()),
            'ShopeePay_Mutation': sum(day['ShopeePay_Mutation'] for day in daily_totals.values() if day['ShopeePay_Mutation'] is not None),
            'ShopeePay_Difference': sum(day['ShopeePay_Difference'] for day in daily_totals.values() if day['ShopeePay_Difference'] is not None),
            'Tiktok_Net': sum(day['Tiktok_Net'] for day in daily_totals.values()),
            'Tiktok_Gross': sum(day['Tiktok_Gross'] for day in daily_totals.values()),
            'Cash_Income': float(cash_income),
            'Cash_Expense': float(cash_expense),
            'Cash_Difference': float(cash_income - cash_expense)
        }

        # Update grand total row to include ShopeePay and mutation data, but remove Grab Mutation and Grab Difference
        dataset.append([
            'Grand Total',
            grand_totals['Gojek_Net'],
            grand_totals['Gojek_Mutation'],
            grand_totals['Gojek_Difference'],
            grand_totals['Grab_Net'],
            grand_totals['Grab_Commission'],
            grand_totals['Shopee_Net'],
            grand_totals['Shopee_Mutation'],
            grand_totals['Shopee_Difference'],
            grand_totals['ShopeePay_Net'],
            grand_totals['ShopeePay_Mutation'],
            grand_totals['ShopeePay_Difference'],
            grand_totals['Tiktok_Net'],
            grand_totals['Cash_Income'],
            grand_totals['Cash_Expense'],
            grand_totals['Cash_Difference']
        ])

        # Create a new workbook and select the active sheet
        wb = Workbook()
        daily_sheet = wb.active
        daily_sheet.title = 'Daily'
        summary_sheet = wb.create_sheet(title='Summary')

        # Style configurations
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        center_align = Alignment(horizontal='center')

        # Daily Sheet Formatting
        # Add title rows
        daily_sheet['A1'] = 'Sales Report'
        daily_sheet['A2'] = 'Period:'
        daily_sheet['B2'] = f'{start_date.strftime("%Y-%m-%d")} to {end_date.strftime("%Y-%m-%d")}'
        daily_sheet['A3'] = 'Outlet:'
        daily_sheet['B3'] = outlet.outlet_name_gojek

        # Define platform colors
        date_fill = PatternFill(start_color='C9F0FF', end_color='C9F0FF', fill_type='solid')  # Light green
        gojek_fill = PatternFill(start_color='00AA13', end_color='00AA13', fill_type='solid')  # Light green
        grab_fill = PatternFill(start_color='98FB98', end_color='98FB98', fill_type='solid')   # Pale green
        shopee_fill = PatternFill(start_color='FF7A00', end_color='FF7A00', fill_type='solid') # Light pink
        cash_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')   # Light blue
        commision_fill = PatternFill(start_color='C6CCB2', end_color='C6CCB2', fill_type='solid') # Light pink
        shopeepay_fill = PatternFill(start_color='E31F26', end_color='E31F26', fill_type='solid')  # Shopee red

        # New: Colors for mutation and difference columns
        mutation_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')   # Light yellow
        difference_fill = PatternFill(start_color='F4CCCC', end_color='F4CCCC', fill_type='solid') # Light red

        # Headers with their corresponding colors
        header_colors = {
            'Date': date_fill,
            'Gojek Net': gojek_fill,
            'Gojek Mutation': gojek_fill, 'Gojek Difference': difference_fill,
            'Grab Net': grab_fill,
            'Grab Mutation': grab_fill, 'Grab Difference': difference_fill,
            'Shopee Net': shopee_fill,
            'Shopee Mutation': shopee_fill, 'Shopee Difference': difference_fill,
            'ShopeePay Net': shopeepay_fill,
            'ShopeePay Mutation': shopeepay_fill, 'ShopeePay Difference': difference_fill,
            'Tiktok Net': PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid'),  # Light pink for TikTok
            'Cash Income': cash_fill, 'Cash Expense': cash_fill, 'Sisa Cash (Admin)': cash_fill
        }

        headers = [
            'Date',
            'Gojek Net', 'Gojek Mutation', 'Gojek Difference',
            'Grab Net',
            'Grab Net (after commission)',
            'Shopee Net', 'Shopee Mutation', 'Shopee Difference',
            'ShopeePay Net','ShopeePay Mutation', 'ShopeePay Difference', 'Tiktok Net', 'Cash Income (Admin)', 'Cash Expense (Admin)', 'Sisa Cash (Admin)','Minusan (Mutasi)'
        ]

        # Write headers to Excel
        for col, header in enumerate(headers, 1):
            cell = daily_sheet.cell(row=5, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_colors.get(header, header_fill)
            cell.alignment = center_align

        # Write daily data rows to Excel using all_dates
        current_row = 6
        for date in all_dates:
            totals = daily_totals[date]
            minusan_total = minusan_by_date.get(date, 0)
            daily_commission = (totals['Grab_Net'] - (totals['Grab_Net'] * 1/74)) if outlet.brand not in ["Pukis & Martabak Kota Baru"] else 0
            totals['Grab_Commission'] = daily_commission            
            row_data = [
                date,
                totals['Gojek_Net'],
                totals['Gojek_Mutation'],
                totals['Gojek_Difference'],
                totals['Grab_Net'],
                totals['Grab_Commission'],
                totals['Shopee_Net'],
                totals['Shopee_Mutation'],
                totals['Shopee_Difference'],
                totals['ShopeePay_Net'],
                totals['ShopeePay_Mutation'],
                totals['ShopeePay_Difference'],
                totals['Tiktok_Net'],   
                totals['Cash_Income'],
                totals['Cash_Expense'],
                totals['Cash_Income'] - totals['Cash_Expense'],
                minusan_total
            ]
            for col, value in enumerate(row_data, 1):
                cell = daily_sheet.cell(row=current_row, column=col)
                cell.value = value
                if col > 1:  # Format numbers
                    cell.number_format = '#,##0'

                # Apply color for Difference columns (Gojek, Shopee)
                # Columns: 4 (Gojek Difference), 10 (Shopee Difference)
                if col in [4, 10, 13]:
                    if value is not None:
                        if value > 0:
                            cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Light green
                        elif value < 0:
                            cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # Light red
                        else:
                            cell.fill = difference_fill  # Neutral color
                    else:
                        cell.fill = difference_fill  # Neutral color

            current_row += 1

        # Update grand total row to match new headers
        grand_total_row = current_row
        daily_sheet.cell(row=grand_total_row, column=1, value='Grand Total').font = header_font
        grand_total_data = [
            grand_totals['Gojek_Net'],
            grand_totals['Gojek_Mutation'],
            grand_totals['Gojek_Difference'],
            grand_totals['Grab_Net'],
            # grand_totals['Grab_Mutation'],
            # grand_totals['Grab_Difference'],
            grand_totals['Grab_Commission'],
            grand_totals['Shopee_Net'],
            grand_totals['Shopee_Mutation'],
            grand_totals['Shopee_Difference'],
            grand_totals['ShopeePay_Net'],
            grand_totals['ShopeePay_Mutation'],
            grand_totals['ShopeePay_Difference'],
            grand_totals['Tiktok_Net'],
            grand_totals['Cash_Income'],
            grand_totals['Cash_Expense'],
            grand_totals['Cash_Difference']
        ]
        for col, value in enumerate(grand_total_data, 2):
            cell = daily_sheet.cell(row=grand_total_row, column=col)
            cell.value = value
            cell.font = header_font
            cell.fill = yellow_fill
            cell.alignment = Alignment(horizontal='center', vertical='center') 
            cell.number_format = '#,##0'

                
        #Pukis Sheet Formatting
        #Exclusive for outlet.brand = "Pukis & Martabak Kota Baru"
        if outlet.brand == "Pukis & Martabak Kota Baru":
            pukis_reports = Pukis.query.filter(
                Pukis.outlet_code == outlet_code,
                Pukis.tanggal >= start_date,
                Pukis.tanggal <= end_date_inclusive
            ).order_by(Pukis.tanggal.asc()).all()

            pukis_sheet = wb.create_sheet(title='Pukis Inventory')
            pukis_sheet['A1'] = 'Pukis Daily Inventory'
            pukis_sheet['A2'] = 'Period:'
            pukis_sheet['B2'] = f'{start_date.strftime("%Y-%m-%d")} to {end_date.strftime("%Y-%m-%d")}'
            pukis_sheet['A3'] = 'Outlet:'
            pukis_sheet['B3'] = outlet.outlet_name_gojek
            headers = [
                'Date', 'Inventory Type', 'Product Type', 'Amount'
            ]
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # Light yellow
            data_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')  # White
            sisa_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Light green
            # Write headers with style
            for col, header in enumerate(headers, 1):
                cell = pukis_sheet.cell(row=5, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')

            # Group pukis_reports by date and product_type
            pukis_by_date_type = defaultdict(lambda: defaultdict(dict))
            for report in pukis_reports:
                date_str = report.tanggal.strftime('%Y-%m-%d')
                product_type = report.pukis_product_type
                pukis_by_date_type[date_str][product_type][report.pukis_inventory_type] = float(report.amount or 0)

            current_row = 6
            for date_str in sorted(pukis_by_date_type.keys()):
                for product_type in ['jumbo', 'klasik']:
                    daily = pukis_by_date_type[date_str][product_type]
                    # Write the four inventory rows (produksi, terjual, retur, free)
                    for inv_type in ['produksi', 'terjual', 'retur', 'free']:
                        amount = daily.get(inv_type, 0)
                        cell_date = pukis_sheet.cell(row=current_row, column=1, value=date_str)
                        cell_date.number_format = 'yyyy-mm-dd'
                        cell_date.alignment = Alignment(horizontal='center')
                        cell_type = pukis_sheet.cell(row=current_row, column=2, value=inv_type.capitalize())
                        cell_type.alignment = Alignment(horizontal='center')
                        cell_prod = pukis_sheet.cell(row=current_row, column=3, value=product_type.capitalize())
                        cell_prod.alignment = Alignment(horizontal='center')
                        cell_amt = pukis_sheet.cell(row=current_row, column=4, value=amount)
                        cell_amt.number_format = '#,##0'
                        cell_amt.alignment = Alignment(horizontal='right')
                        # Fill
                        for col in range(1, 5):
                            pukis_sheet.cell(row=current_row, column=col).fill = data_fill
                        current_row += 1
                    # Calculate and write the custom "sisa" row
                    produksi = daily.get('produksi', 0)
                    terjual = daily.get('terjual', 0)
                    retur = daily.get('retur', 0)
                    free = daily.get('free', 0)
                    sisa = produksi - (terjual + retur + free)
                    cell_date = pukis_sheet.cell(row=current_row, column=1, value=date_str)
                    cell_date.number_format = 'yyyy-mm-dd'
                    cell_date.alignment = Alignment(horizontal='center')
                    cell_type = pukis_sheet.cell(row=current_row, column=2, value='Sisa')
                    cell_type.alignment = Alignment(horizontal='center')
                    cell_prod = pukis_sheet.cell(row=current_row, column=3, value=product_type.capitalize())
                    cell_prod.alignment = Alignment(horizontal='center')
                    cell_amt = pukis_sheet.cell(row=current_row, column=4, value=sisa)
                    cell_amt.number_format = '#,##0'
                    cell_amt.alignment = Alignment(horizontal='right')
                    # Sisa row fill
                    for col in range(1, 5):
                        pukis_sheet.cell(row=current_row, column=col).fill = sisa_fill
                    current_row += 1
            # Auto-adjust column widths
            for column_cells in pukis_sheet.columns:
                max_length = 0
                column_cells = list(column_cells)
                for cell in column_cells:
                    try:
                        if cell.value is not None and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                pukis_sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width

            # After writing all daily rows, calculate totals for the date range
            totals = {
                'jumbo': {'produksi': 0, 'terjual': 0, 'retur': 0, 'free': 0},
                'klasik': {'produksi': 0, 'terjual': 0, 'retur': 0, 'free': 0}
            }

            for report in pukis_reports:
                ptype = report.pukis_product_type
                itype = report.pukis_inventory_type
                if ptype in totals and itype in totals[ptype]:
                    totals[ptype][itype] += float(report.amount or 0)

            # Write the totals at the end of the sheet
            summary_start_row = current_row + 2  # Leave a blank row after the last data row

            for product_type in ['jumbo', 'klasik']:
                for inv_type in ['produksi', 'terjual', 'retur', 'free']:
                    pukis_sheet.cell(row=summary_start_row, column=1, value='TOTAL')
                    pukis_sheet.cell(row=summary_start_row, column=2, value=inv_type.capitalize())
                    pukis_sheet.cell(row=summary_start_row, column=3, value=product_type.capitalize())
                    pukis_sheet.cell(row=summary_start_row, column=4, value=totals[product_type][inv_type])
                    summary_start_row += 1
       

        # Summary Sheet Formatting
        # Add titles
        summary_sheet['A1'] = 'Summary Report'
        summary_sheet['A2'] = 'Period:'
        summary_sheet['B2'] = f'{start_date.strftime("%Y-%m-%d")} to {end_date.strftime("%Y-%m-%d")}'
        summary_sheet['A3'] = 'Outlet:'
        summary_sheet['B3'] = outlet.outlet_name_gojek

        # Online Platform Summary
        current_row = 5
        summary_title_cell = summary_sheet.cell(row=current_row, column=1, value='Online Platform Summary')
        summary_title_cell.font = header_font
        summary_title_cell.fill = grab_fill
        current_row += 1
        
        platform_headers = ['Platform', 'Gross', 'Net', 'Difference']
        for col, header in enumerate(platform_headers, 1):
            cell = summary_sheet.cell(row=current_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = grab_fill
        current_row += 1

        # Add platform data
        platforms = [
            ('Gojek', 'Gojek_Gross', 'Gojek_Net',  'Gojek_Mutation'),
            ('Grab (Total)', 'Grab_Gross', 'Grab_Net', None
            #  'Grab_Mutation'
            ),
            ('   GrabFood', grabfood_gross_total, grabfood_net_total, None),  # Fixed: use grabfood_net_total
            ('   OVO', grabovo_gross_total, grabovo_net_total, None),         # Already correct
            ('Shopee', 'Shopee_Gross', 'Shopee_Net', 'Shopee_Mutation'),
            ('ShopeePay', 'ShopeePay_Gross', 'ShopeePay_Net', 'ShopeePay_Mutation'),
            ('Tiktok', 'Tiktok_Gross', 'Tiktok_Net', None)  # Tiktok does not have mutation data
        ]
        
        for platform_data in platforms:
            if isinstance(platform_data[1], (int, float)):  # For GrabFood and GrabOVO
                platform, gross, net, _ = platform_data
                row_data = [
                    platform,
                    gross,
                    net,
                    gross - net
                ]
            else:
                platform, gross_key, net_key, mutation_key = platform_data
                gross = grand_totals[gross_key]
                # Use mutation total if available, else use net total
                # if mutation_key and grand_totals.get(mutation_key) is not None and grand_totals.get(mutation_key) != 0:
                #     net = grand_totals[mutation_key]
                # else:
                net = grand_totals[net_key]
                row_data = [
                    platform,
                    gross,
                    net,
                    gross - net
                ]
            
            for col, value in enumerate(row_data, 1):
                cell = summary_sheet.cell(row=current_row, column=col)
                cell.value = value
                cell.fill = grab_fill
                if col > 1:
                    cell.number_format = '#,##0'
            current_row += 1
            
       

        current_row += 1  # Add spacing

        # Income/Expense Summary
        summary_title_cell = summary_sheet.cell(row=current_row, column=1, value='Income/Expense Summary')
        summary_title_cell.font = header_font
        summary_title_cell.fill = cash_fill
        current_row += 1
        
        expense_headers = ['Category', 'Income', 'Expense', 'Net Total', 'Description', 'Date Range']
        for col, header in enumerate(expense_headers, 1):
            cell = summary_sheet.cell(row=current_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = cash_fill
        current_row += 1

        # Query manual entries with joined category (income or expense)
        from sqlalchemy.orm import aliased
        from app.models.income_category import IncomeCategory
        from app.models.expense_category import ExpenseCategory
        IncomeCat = aliased(IncomeCategory)
        ExpenseCat = aliased(ExpenseCategory)
        manual_entries = (
            ManualEntry.query
            .filter(
                ManualEntry.outlet_code == outlet_code,
                ManualEntry.start_date <= end_date,
                ManualEntry.end_date >= start_date,
                ~ManualEntry.description.ilike('%minus%')
            )
            .outerjoin(IncomeCat, (ManualEntry.category_id == IncomeCat.id) & (ManualEntry.entry_type == 'income'))
            .outerjoin(ExpenseCat, (ManualEntry.category_id == ExpenseCat.id) & (ManualEntry.entry_type == 'expense'))
            .add_entity(IncomeCat)
            .add_entity(ExpenseCat)
            .all()
        )

        # Calculate manual entry totals
        manual_income = sum(float(entry.amount or 0) for entry, income_cat, expense_cat in manual_entries if entry.entry_type == 'income')
        manual_expense = sum(float(entry.amount or 0) for entry, income_cat, expense_cat in manual_entries if entry.entry_type == 'expense')

        # Add cash data
        cash_net_total = grand_totals['Cash_Income'] - grand_totals['Cash_Expense']
        cash_row = ['Cash', grand_totals['Cash_Income'], grand_totals['Cash_Expense'], cash_net_total, '', '']
        for col, value in enumerate(cash_row, 1):
            cell = summary_sheet.cell(row=current_row, column=col)
            cell.value = value
            cell.fill = cash_fill
            if 1 < col <= 4:  # Format numbers for Income, Expense, and Net Total columns
                cell.number_format = '#,##0'
        current_row += 1

        # Sort manual entries by date parsed from Indonesian month names in description
        import re
        MONTH_MAP = {
            'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'Mei': 5, 'Jun': 6,
            'Jul': 7, 'Agu': 8, 'Sep': 9, 'Okt': 10, 'Nov': 11, 'Des': 12
        }
        def parse_indonesian_date(description):
            match = re.search(r'(\d{1,2})\s+([A-Za-z]+)', description)
            if match:
                day = int(match.group(1))
                month_str = match.group(2).capitalize()
                month = MONTH_MAP.get(month_str)
                if month:
                    year = datetime.now().year
                    return datetime(year, month, day)
            return datetime.min

        manual_entries_sorted = sorted(
            manual_entries,
            key=lambda x: parse_indonesian_date(x[0].description)
        )

        # Add manual entries with their details (category name from joined table)
        for entry, income_cat, expense_cat in manual_entries_sorted:
            amount = float(entry.amount or 0)
            income_amount = amount if entry.entry_type == 'income' else 0
            expense_amount = amount if entry.entry_type == 'expense' else 0
            net_amount = income_amount - expense_amount
            if entry.entry_type == 'income':
                category_name = income_cat.name if income_cat else ''
            else:
                category_name = expense_cat.name if expense_cat else ''
            row_data = [
                category_name,
                income_amount,
                expense_amount,
                net_amount,
                entry.description,
                f"{entry.start_date} - {entry.end_date}"
            ]
            for col, value in enumerate(row_data, 1):
                cell = summary_sheet.cell(row=current_row, column=col)
                cell.value = value
                cell.fill = cash_fill
                if 1 < col <= 4:  # Format numbers for Income, Expense, and Net Total columns
                    cell.number_format = '#,##0'
            current_row += 1

        # Add total row
        total_net = (grand_totals['Cash_Income'] + manual_income) - (grand_totals['Cash_Expense'] + manual_expense)
        total_row = [
            'Total',
            grand_totals['Cash_Income'] + manual_income,
            grand_totals['Cash_Expense'] + manual_expense,
            total_net,
            '',
            ''
        ]
        for col, value in enumerate(total_row, 1):
            cell = summary_sheet.cell(row=current_row, column=col)
            cell.value = value
            cell.font = header_font
            cell.fill = cash_fill
            if 1 < col <= 4:  # Format numbers for Income, Expense, and Net Total columns
                cell.number_format = '#,##0'
        current_row += 1

        current_row += 1  # Add spacing

        # Calculate total net income from all sources
        online_net_total = (
            grand_totals['Gojek_Net'] +
            grand_totals['Grab_Net'] +
            grand_totals['Shopee_Net'] +
            grand_totals['ShopeePay_Net'] + grand_totals['Tiktok_Net']
        )
        
        # Calculate commission first (moved from below)
        management_commission = grabfood_net_total * 1/74 if outlet.brand in ["MP78", "MP78 Express", "Martabak 777 Sinar Bulan","Martabak 999 Asli Bandung", "Martabak Surya Kencana","Martabak Akim"] else 0
        partner_commission = grabfood_net_total * 1/74 if outlet.brand in ["MP78", "MP78 Express", "Martabak 777 Sinar Bulan","Martabak 999 Asli Bandung", "Martabak Surya Kencana","Martabak Akim"] else 0
        tiktok_commission = grand_totals['Tiktok_Net'] * 1/74 if outlet.brand in ["MP78", "MP78 Express", "Martabak 777 Sinar Bulan","Martabak 999 Asli Bandung", "Martabak Surya Kencana","Martabak Akim"] else 0

        cash_manual_net_total = (
            (grand_totals['Cash_Income'] + manual_income) -
            (grand_totals['Cash_Expense'] + manual_expense)
        )
        
        # Calculate final online net total after commission
        online_net_after_commission = online_net_total - management_commission

        summary_data = [('Online Platforms Net', online_net_total)]
        
        # Add commission-related entries only for specific brands
        if outlet.brand in ["MP78", "MP78 Express", "Martabak 777 Sinar Bulan","Martabak 999 Asli Bandung", "Martabak Surya Kencana","Martabak Akim"]:
            summary_data.extend([
                ('Grab Management', -management_commission),
                ('Online Net After Commission', online_net_after_commission)
            ])
        
        summary_data.extend([
            ('Cash & Manual Net', cash_manual_net_total),
            ('GRAND TOTAL NET INCOME', online_net_after_commission + cash_manual_net_total)
        ])

        for label, amount in summary_data:
            cell = summary_sheet.cell(row=current_row, column=1, value=label)
            amount_cell = summary_sheet.cell(row=current_row, column=2, value=amount)
            cell.fill = shopee_fill
            amount_cell.number_format = '#,##0'
            amount_cell.fill = shopee_fill
            if label == 'GRAND TOTAL NET INCOME':
                cell.font = header_font
                amount_cell.font = header_font
                cell.fill = yellow_fill
                amount_cell.fill = yellow_fill
            elif label == 'Management Commission':
                cell.font = Font(bold=True, color='FF0000')  # Red color for commission
                amount_cell.font = Font(bold=True, color='FF0000')
            current_row += 1

        current_row += 2  # Add spacing


        # Commission Summary - Only show for specific brands
        if outlet.brand in ["MP78", "MP78 Express", "Martabak 777 Sinar Bulan","Martabak 999 Asli Bandung", "Martabak Surya Kencana","Martabak Akim"]:
            summary_title_cell = summary_sheet.cell(row=current_row, column=1, value='Commission Summary')
            summary_title_cell.font = header_font
            summary_title_cell.fill = commision_fill
            current_row += 1
            
            commission_headers = ['Category','Rate', 'Commission']
            for col, header in enumerate(commission_headers, 1):
                cell = summary_sheet.cell(row=current_row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = commision_fill
            current_row += 1

            commission_data = [
                ('Management Commission (GrabFood)', '1%', management_commission),
                ('Partner Commission (GrabFood)', '1%', partner_commission),
                ('Management Commission (TikTok)', '1%', tiktok_commission),

            ]

            for category, rate, commission in commission_data:
                row_data = [category, rate, commission]
                for col, value in enumerate(row_data, 1):
                    cell = summary_sheet.cell(row=current_row, column=col)
                    cell.value = value
                    cell.fill = commision_fill
                    if col in [3]:  # Format numbers for Base Amount and Commission columns
                        cell.number_format = '#,##0'
                current_row += 1

            current_row += 1  # Add spacing
       

        # Auto-adjust column widths
        for sheet in [daily_sheet, summary_sheet]:
            for column in sheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

       

    #         # Protect each worksheet
    #         sheet.protection.sheet = True
    #         sheet.protection.enable()
    #    # Protect workbook structure properly
    #     wb.security = WorkbookProtection(workbookPassword=None, lockStructure=True)
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

      

        if outlet.brand == "Pukis & Martabak Kota Baru":
        # Split name into brand and location parts
            parts = outlet.outlet_name_gojek.split(',')
            
            # Create short brand code
            brand_code = 'PKB'  # Abbreviation for "Pukis & Martabak Kota Baru"
            
            # Extract location and format it
            location = parts[1].strip().replace(' ', '') if len(parts) > 1 else 'Unknown'

            # Combine brand and location
            safe_outlet_name = f"{brand_code}_{location}"

            # Build filename
            filename = f"Report_{safe_outlet_name}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
        elif outlet.brand == "MP78":
            location_clean = "Unknown"
            if ',' in outlet.outlet_name_gojek:
                _, location = outlet.outlet_name_gojek.split(',', 1)
                location_clean = location.strip().replace(' ', '')

            safe_outlet_name = f"MP78_{location_clean}"
            filename = f"Report_{safe_outlet_name}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
        else:
            safe_outlet_name = outlet.outlet_name_gojek.replace('/', '_').replace('\\', '_').replace(' ', '_')
            filename = f"Report_{safe_outlet_name}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        return jsonify({
            "error": "General process failed",
            "details": str(e),
            "type": type(e).__name__,
            "step": "general"
        }), 400
