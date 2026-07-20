from io import BytesIO
from fpdf import FPDF
from flask import Blueprint, jsonify, request, Response, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from app.models.gojek_reports import GojekReport
from app.models.shopee_reports import ShopeeReport
from app.models.grabfood_reports import GrabFoodReport
from app.models.cash_reports import CashReport
from app.models.manual_entry import ManualEntry
from app.models.shopeepay_reports import ShopeepayReport
from app.models.bank_mutations import BankMutation
from app.models.mp78_mutations import MP78Mutation
from app.models.expense_category import ExpenseCategory
from app.models.income_category import IncomeCategory
from app.models.pukis import Pukis
from app.models.tiktok_reports import TiktokReport
from app.models.outlet_count_pkb import OutletCountPKB
from app.models.ultra_voucher import VoucherReport
from app.models.webshop_report import WebshopReport
from app.extensions import db
import sys
from app.services.consolidation_service import update_daily_total_for_outlet

import calendar
from datetime import datetime
from flask_cors import cross_origin

def parse_date(date_str, default_year=None):
    date_str = date_str.strip()
    formats = [
        '%d-%b-%y',
        '%m/%d/%Y',
        '%d/%m/%Y',
        '%Y-%m-%d',
        '%d-%b'  # Add this format for '10-Apr'
    ]
    for fmt in formats:
        try:
            if fmt == '%d-%b':
                # If no year, use current year or provided default_year
                dt = datetime.strptime(date_str, fmt)
                year = default_year or datetime.now().year
                return dt.replace(year=year)
            return datetime.strptime(date_str, fmt)
        except ValueError:
            pass
    raise ValueError(f"Unknown date format: {date_str}")


def parse_date_range(args):
    start_date_param = args.get("start_date")
    end_date_param = args.get("end_date")

    if not start_date_param and not end_date_param:
        return None, None, None
    if not start_date_param or not end_date_param:
        return None, None, "start_date and end_date are required together"

    try:
        start_date = datetime.strptime(start_date_param, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_date_param, "%Y-%m-%d").date()
    except ValueError:
        return None, None, "start_date and end_date must be in YYYY-MM-DD format"

    if start_date > end_date:
        return None, None, "start_date must be less than or equal to end_date"

    return start_date, end_date, None


def parse_date_range_from_body(data):
    start_date_param = data.get("start_date")
    end_date_param = data.get("end_date")

    if not start_date_param or not end_date_param:
        return None, None, "start_date and end_date are required"

    try:
        start_date = datetime.strptime(start_date_param, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_date_param, "%Y-%m-%d").date()
    except (TypeError, ValueError):
        return None, None, "start_date and end_date must be in YYYY-MM-DD format"

    if start_date > end_date:
        return None, None, "start_date must be less than or equal to end_date"

    return start_date, end_date, None


def _month_last_day(year, month):
    return calendar.monthrange(year, month)[1]


def _add_month(year, month):
    if month == 12:
        return year + 1, 1
    return year, month + 1


def _previous_month(year, month):
    if month == 1:
        return year - 1, 12
    return year, month - 1


def _date_with_clamped_day(year, month, day):
    return datetime(year, month, min(day, _month_last_day(year, month))).date()


def _parse_closing_date_range(closing_date):
    if not closing_date or '-' not in str(closing_date):
        return None

    start_day_str, end_day_str = str(closing_date).split('-', 1)
    try:
        start_day = int(start_day_str.strip())
        end_day = int(end_day_str.strip())
    except ValueError:
        return None

    if not (1 <= start_day <= 31 and 1 <= end_day <= 31):
        return None

    return start_day, end_day


def resolve_upload_manual_entry_date_range(outlet, entry_date):
    if hasattr(entry_date, 'date'):
        entry_date = entry_date.date()

    closing_range = _parse_closing_date_range(
        outlet.closing_date if outlet else None
    )
    if not closing_range:
        return entry_date, entry_date

    start_day, end_day = closing_range
    start_year, start_month = _previous_month(entry_date.year, entry_date.month)
    start_date = _date_with_clamped_day(start_year, start_month, start_day)
    end_date = _date_with_clamped_day(entry_date.year, entry_date.month, end_day)

    if start_date <= entry_date <= end_date:
        return start_date, end_date

    next_end_year, next_end_month = _add_month(entry_date.year, entry_date.month)
    next_start_date = _date_with_clamped_day(entry_date.year, entry_date.month, start_day)
    next_end_date = _date_with_clamped_day(next_end_year, next_end_month, end_day)
    return next_start_date, next_end_date


# import pdfkit
import csv
from io import StringIO
from sqlalchemy import func, or_
from datetime import datetime
from sqlalchemy.exc import IntegrityError
from app.models.outlet import Outlet
import io
from datetime import datetime

import openpyxl
from flask import Blueprint, jsonify, request, send_file
from openpyxl.workbook import Workbook

from app.services.excel_export.sheets.monthly_income_sheet import MonthlyIncomeSheet
from app.services.excel_export.sheets.monthly_management_commission_sheet import (
    MonthlyManagementCommissionSheet,
)
from app.services.excel_export.sheets.monthly_mpr_commission_sheet import (
    MonthlyMprCommissionSheet,
)
from app.services.reporting_service import (
    generate_monthly_management_commission_data_custom_range,
    generate_monthly_management_commission_data,
    generate_monthly_mpr_commission_data,
    generate_monthly_net_income_data,
)
from app.services.mpr_totals_service import (
    calculate_mpr_totals,
    get_mpr_mapping_for_outlet,
)
from app.utils.report_generator import generate_daily_report
# from app.services.excel_export.data_service import (
#     get_kas_transactions, create_kas_transaction,
#     update_kas_transaction, delete_kas_transaction
# )
# from app.services.excel_export.sheets.kas_sheet import KasSheet
# from app.models.kas_transaction import KasTransaction

# config = pdfkit.configuration(wkhtmltopdf='C:/Program Files/wkhtmltopdf/bin/wkhtmltopdf.exe')
reports_bp = Blueprint('reports', __name__, url_prefix='/reports')

import os


def update_store_ids_batch(store_id_map, platform):
    """Update store IDs in batches to prevent timeout"""
    batch_size = 100
    updated = 0
    
    # Process in batches
    store_items = list(store_id_map.items())
    for i in range(0, len(store_items), batch_size):
        batch = store_items[i:i + batch_size]
        for store_name, store_id in batch:
            try:
                # For Gojek
                if platform == 'gojek':
                    # Check if store_id already exists
                    existing_outlet = Outlet.query.filter_by(store_id_gojek=store_id).first()
                    if existing_outlet:
                        continue
                    
                    outlet = Outlet.query.filter_by(outlet_name_gojek=store_name).first()
                    if outlet and not outlet.store_id_gojek:
                        outlet.store_id_gojek = store_id
                        updated += 1
                
                # For Grab
                elif platform == 'grab':
                    existing_outlet = Outlet.query.filter_by(store_id_grab=store_id).first()
                    if existing_outlet:
                        continue
                    
                    outlet = Outlet.query.filter_by(outlet_name_grab=store_name).first()
                    if outlet and not outlet.store_id_grab:
                        outlet.store_id_grab = store_id
                        updated += 1

                # For Shopee
                elif platform == 'shopee':
                    existing_outlet = Outlet.query.filter_by(store_id_shopee=store_id).first()
                    if existing_outlet:
                        continue
                    
                    outlet = Outlet.query.filter_by(outlet_name_grab=store_name).first()
                    if outlet and not outlet.store_id_shopee:
                        outlet.store_id_shopee = store_id
                        updated += 1
                
            except Exception as e:
                continue
        
        # Commit each batch
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()

    return updated

@reports_bp.route('/upload/shopee_adjustment', methods=['POST'])
def upload_report_shopee_adjustment():
    files = request.files.getlist('file')
    if not files:
        files = [request.files.get('file')]
    
    if not files or not files[0]:
        return jsonify({'msg': 'No files uploaded'}), 400

    try:
        total_reports = 0
        skipped_reports = 0
        store_id_map = {}
        
        for file in files:
            file_contents = file.read().decode('utf-8')
            csv_file = StringIO(file_contents)
            reader = csv.DictReader(csv_file)
            
            reports = []
            for row in reader:
                store_name = row.get('Store Name', '').strip()
                store_id = row.get('Store ID')
                if store_name and store_id:
                    store_id_map[store_name] = store_id

                outlet = None
                if store_id:
                    outlet = Outlet.query.filter_by(store_id_shopee=store_id).first()
                if not outlet and store_name:
                    outlet = Outlet.query.filter_by(outlet_name_grab=store_name).first()
                if not outlet:
                    print(f"SKIPPED: Outlet not found for Store Name '{store_name}', Store ID '{store_id}' | Row: {row}")
                    skipped_reports += 1
                    continue

                # Check if adjustment already exists
                refund_id = row.get('Wallet Adjustment ID', '')
                adjustment_time_str = row.get('Wallet Adjustment Time', '')
                adjustment_time = None
                if adjustment_time_str:
                    try:
                        adjustment_time = datetime.strptime(adjustment_time_str, '%Y-%m-%d %H:%M:%S')
                    except Exception as e:
                        print(f"SKIPPED: Invalid adjustment time format: {adjustment_time_str} | Row: {row}")
                        skipped_reports += 1
                        continue

                if refund_id and adjustment_time:
                    existing_report = ShopeeReport.query.filter_by(
                        order_id=refund_id,
                        transaction_type='Adjustment',
                        order_create_time=adjustment_time
                    ).first()
                    if existing_report:
                        skipped_reports += 1
                        continue

                try:
                    # Parse the adjustment time
                    adjustment_time = datetime.strptime(row.get('Wallet Adjustment Time', ''), '%Y-%m-%d %H:%M:%S')

                    report = ShopeeReport(
                        brand_name=outlet.brand,
                        outlet_code=outlet.outlet_code,
                        transaction_type='Adjustment',  # Mark as adjustment
                        order_id=row.get('Wallet Adjustment ID', ''),  # Use refund ID as order ID
                        store_id=store_id,
                        store_name=store_name,
                        order_create_time=adjustment_time,
                        order_amount=float(row.get('Wallet Adjustment Amount', 0) or 0),
                        total=float(row.get('Wallet Adjustment Amount', 0) or 0),
                        net_income=float(row.get('Wallet Adjustment Amount', 0) or 0),
                        order_status='Completed',
                        order_type=row.get('Wallet Adjustment Reason', 'Adjustment')
                    )
                    reports.append(report)
                    total_reports += 1
                except (ValueError, TypeError) as e:
                    skipped_reports += 1
                    continue

            if reports:
                try:
                    db.session.bulk_save_objects(reports)
                    db.session.commit()
                except IntegrityError:
                    db.session.rollback()
                    for report in reports:
                        try:
                            db.session.add(report)
                            db.session.commit()
                        except IntegrityError:
                            db.session.rollback()
                            skipped_reports += 1
                        except Exception as e:
                            db.session.rollback()
                            print(f"Error saving report: {str(e)}")

        updated_count = update_store_ids_batch(store_id_map, 'shopee')
        
        return jsonify({
            'msg': 'Adjustment reports uploaded successfully',
            'total_records': total_reports,
            'skipped_records': skipped_reports,
            'store_ids_updated': updated_count
        }), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@reports_bp.route('/upload/voucher', methods=['POST'])
def upload_voucher_report():
    """
    Handles the upload and processing of voucher report CSV files.
    """
    files = request.files.getlist('file')
    if not files or not files[0]:
        return jsonify({'msg': 'No files uploaded'}), 400

    try:
        total_reports_processed = 0
        skipped_reports = 0
        
        for file in files:
            if file.filename == '':
                continue

            file_contents = file.read().decode('utf-8-sig') # Use 'utf-8-sig' to handle potential BOM
            csv_file = StringIO(file_contents)
            reader = csv.DictReader(csv_file)
            
            reports_to_add = []
            for row in reader:
                # 1. Parse the row using the static method in the VoucherReport model
                parsed_data = VoucherReport.parse_row(row)

                if not parsed_data:
                    # parse_row already prints the error, so we just count and skip
                    skipped_reports += 1
                    continue

                # 2. Check for duplicates before adding
                order_no = parsed_data.get('order_no')
                if not order_no:
                    print(f"SKIPPED: 'Order No' is missing. | Row: {row}")
                    skipped_reports += 1
                    continue

                existing_report = VoucherReport.query.filter_by(order_no=order_no).first()
                if existing_report:
                    print(f"SKIPPED: Duplicate Order No '{order_no}' found.")
                    skipped_reports += 1
                    continue

                # 3. Verify that the outlet_code exists in the Outlet table (optional but good practice)
                outlet_code = parsed_data.get('outlet_code')
                if outlet_code:
                    outlet = Outlet.query.filter_by(outlet_code=outlet_code).first()
                    if not outlet:
                        print(f"SKIPPED: Outlet with code '{outlet_code}' not found in the database. | Row: {row}")
                        skipped_reports += 1
                        continue

                # 4. Create the report object if all checks pass
                try:
                    report = VoucherReport(**parsed_data)
                    reports_to_add.append(report)
                except Exception as e:
                    print(f"SKIPPED: Error creating report object: {e} | Data: {parsed_data}")
                    skipped_reports += 1
                    continue

            # 5. Bulk insert the reports for the current file
            if reports_to_add:
                try:
                    db.session.bulk_save_objects(reports_to_add)
                    db.session.commit()
                    total_reports_processed += len(reports_to_add)
                except IntegrityError:
                    # Fallback to one-by-one insert if bulk fails (e.g., duplicate within the file)
                    db.session.rollback()
                    print("Bulk insert failed, falling back to individual inserts.")
                    for report in reports_to_add:
                        try:
                            # Re-check for duplicates that might exist from the same file batch
                            if not VoucherReport.query.filter_by(order_no=report.order_no).first():
                                db.session.add(report)
                                db.session.commit()
                                total_reports_processed += 1
                            else:
                                print(f"SKIPPED (Fallback): Duplicate Order No '{report.order_no}'.")
                                skipped_reports +=1
                        except Exception as e:
                            db.session.rollback()
                            print(f"SKIPPED (Fallback): Error saving report for Order No '{report.order_no}': {e}")
                            skipped_reports += 1
                except Exception as e:
                    db.session.rollback()


        return jsonify({
            'msg': 'Voucher reports uploaded successfully.',
            'total_records_processed': total_reports_processed,
            'skipped_records': skipped_reports
        }), 201

    except Exception as e:
        db.session.rollback()
        # It's good practice to log the full error here
        # import traceback
        # traceback.print_exc()
        return jsonify({'error': f'An unexpected error occurred: {str(e)}'}), 500
        

@reports_bp.route('/upload/mutation', methods=['POST'])
def upload_report_mutation():
    files = request.files.getlist('file')
    rekening_number = request.form.get('rekening_number')

    if not files:
        files = [request.files.get('file')]
    if not files or not files[0]:
        return jsonify({'msg': 'No files uploaded'}), 400
    if not rekening_number:
        return jsonify({'msg': 'Rekening number is required'}), 400

    total_mutations = 0
    skipped_mutations = 0
    debug_skipped = []

    def _is_blank_row(row):
        return not row or not any(str(value).strip() for value in row if value is not None)

    def _is_mutation_table_header(row):
        row_text = " ".join(str(value).strip().upper() for value in row if value is not None)
        return (
            ('TANGGAL' in row_text or 'DATE' in row_text)
            and ('KETERANGAN' in row_text or 'DESCRIPTION' in row_text)
            and ('CABANG' in row_text or 'BRANCH' in row_text)
            and ('JUMLAH' in row_text or 'AMOUNT' in row_text)
            and ('SALDO' in row_text or 'BALANCE' in row_text)
        )

    def _debug_row(row):
        return [
            value.isoformat() if hasattr(value, 'isoformat') else value
            for value in row
        ]

    def _mutation_rows_from_excel(file_bytes):
        workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
        rows = []

        for worksheet in workbook.worksheets:
            rows.extend(
                _mutation_rows_from_table(
                    enumerate(worksheet.iter_rows(values_only=True), start=1)
                )
            )

        return rows

    def _mutation_rows_from_table(numbered_rows, min_header_row=1):
        rows = []
        in_mutation_table = False
        header_row = None

        for row_number, row in numbered_rows:
            row = list(row)
            if _is_blank_row(row):
                continue
            if row_number >= min_header_row and _is_mutation_table_header(row):
                in_mutation_table = True
                header_row = row
                continue
            if not in_mutation_table:
                continue

            first_cell = str(row[0]).strip().upper() if row[0] is not None else ''
            if first_cell.startswith(('SALDO', 'MUTASI')):
                break

            rows.append((row_number, header_row, row))

        return rows

    def _mutation_rows_from_upload(file):
        file_bytes = file.read()
        filename = (file.filename or '').lower()
        if filename.endswith(('.xlsx', '.xlsm')):
            return _mutation_rows_from_excel(file_bytes)

        file_contents = file_bytes.decode('utf-8-sig')
        csv_file = StringIO(file_contents)
        reader = csv.reader(csv_file)
        rows = _mutation_rows_from_table(enumerate(reader, start=1))
        if rows:
            return rows

        csv_file.seek(0)
        reader = csv.reader(csv_file)
        next(reader, None)  # Skip legacy CSV header
        return [(idx, None, row) for idx, row in enumerate(reader, start=2)]

    try:
        for file in files:
            mutations = []
            mp78_mutations = []
            for row_number, header_row, row in _mutation_rows_from_upload(file):
                try:
                    # Skip if the date column is 'PEND'
                    first_cell = str(row[0]).strip().upper() if row and row[0] is not None else ''
                    if first_cell == 'PEND':
                        skipped_mutations += 1
                        debug_skipped.append({
                            'row_number': row_number,
                            'reason': "Date column is 'PEND'",
                            'row': _debug_row(row)
                        })
                        continue

                    parsed = BankMutation.parse_mutation_row(row, rekening_number)
                    if not parsed:
                        skipped_mutations += 1
                        debug_skipped.append({
                            'row_number': row_number,
                            'reason': 'Unknown platform or parse failed',
                            'header': _debug_row(header_row) if header_row else None,
                            'row': _debug_row(row)
                        })
                        continue

                    mutation_model = parsed.pop('_mutation_model', 'bank')
                    if mutation_model == 'mp78':
                        mutation = MP78Mutation(
                            rekening_number=rekening_number,
                            **parsed
                        )

                        exists = MP78Mutation.query.filter_by(
                            transaction_id=mutation.transaction_id,
                        ).first()
                        if exists:
                            skipped_mutations += 1
                            debug_skipped.append({
                                'row_number': row_number,
                                'reason': 'Duplicate MP78 mutation entry',
                                'row': _debug_row(row)
                            })
                            continue

                        mp78_mutations.append(mutation)
                        total_mutations += 1
                        continue

                    mutation = BankMutation(
                        rekening_number=rekening_number,
                        **parsed
                    )

                    # Check for duplicates in the database only
                    exists = BankMutation.query.filter_by(
                        tanggal=mutation.tanggal,
                        transaction_amount=mutation.transaction_amount,
                        platform_code=mutation.platform_code,
                    ).first()
                    if exists:
                        skipped_mutations += 1
                        debug_skipped.append({
                            'row_number': row_number,
                            'reason': 'Duplicate mutation entry',
                            'row': _debug_row(row)
                        })
                        continue

                    mutations.append(mutation)
                    total_mutations += 1

                except Exception as e:
                    skipped_mutations += 1
                    debug_skipped.append({
                        'row_number': row_number,
                        'reason': f'Exception: {str(e)}',
                        'row': _debug_row(row)
                    })

            # Bulk save mutations for this file
            if mutations:
                try:
                    db.session.bulk_save_objects(mutations)
                    db.session.commit()
                except IntegrityError:
                    db.session.rollback()
                    # Handle mutations one by one if bulk insert fails
                    for mutation in mutations:
                        try:
                            db.session.add(mutation)
                            db.session.commit()
                        except IntegrityError:
                            db.session.rollback()
                            skipped_mutations += 1
                        except Exception as e:
                            db.session.rollback()
                            debug_skipped.append({
                                'row_number': None,
                                'reason': f'Error saving mutation: {str(e)}',
                                'row': None
                            })

            if mp78_mutations:
                try:
                    db.session.bulk_save_objects(mp78_mutations)
                    db.session.commit()
                except IntegrityError:
                    db.session.rollback()
                    for mutation in mp78_mutations:
                        try:
                            db.session.add(mutation)
                            db.session.commit()
                        except IntegrityError:
                            db.session.rollback()
                            skipped_mutations += 1
                        except Exception as e:
                            db.session.rollback()
                            debug_skipped.append({
                                'row_number': None,
                                'reason': f'Error saving MP78 mutation: {str(e)}',
                                'row': None
                            })

       
        seen_reasons = set()
        one_per_reason = []

        for entry in debug_skipped:
            reason = entry['reason']
            if reason not in seen_reasons:
                one_per_reason.append(entry)
                seen_reasons.add(reason)

        return jsonify({
            'msg': 'Bank mutations uploaded successfully',
            'total_records': total_mutations,
            'skipped_records': skipped_mutations,
            'skipped_rows_debug': one_per_reason
        }), 201


    except Exception as e:
        db.session.rollback()
        return jsonify({
            'error': str(e),
            'skipped_rows_debug': debug_skipped
        }), 500

@reports_bp.route('/upload/gojek', methods=['POST'])
def upload_report_gojek():
    files = request.files.getlist('file')
    if not files:
        files = [request.files.get('file')]
    
    if not files or not files[0]:
        return jsonify({'msg': 'No files uploaded'}), 400

    try:
        total_reports = 0
        skipped_reports = 0
        store_id_map = {}
        affected_outlets = set()
        seen_transaction_ids = set()
        seen_order_ids = set()
        seen_amount_keys = set()

        outlets = Outlet.query.all()
        outlets_by_store_id = {outlet.store_id_gojek: outlet for outlet in outlets if outlet.store_id_gojek}
        outlets_by_name = {outlet.outlet_name_gojek: outlet for outlet in outlets if outlet.outlet_name_gojek}

        def clean_identifier(value):
            if value is None:
                return None
            value = str(value).strip().strip("'")
            return value or None

        def safe_float(value):
            if not value:
                return 0
            try:
                return float(str(value).replace(',', '').replace("'", ''))
            except (ValueError, TypeError):
                return 0

        def safe_float_or_none(value):
            try:
                return float(str(value).replace(',', '').replace("'", ''))
            except (ValueError, TypeError):
                return None

        def chunks(values, size=1000):
            values = list(values)
            for index in range(0, len(values), size):
                yield values[index:index + size]

        def load_existing_gojek_identifiers(transaction_ids, order_ids):
            existing_transaction_ids = set()
            existing_order_ids = set()

            for batch in chunks(transaction_ids):
                existing_transaction_ids.update(
                    identifier
                    for (identifier,) in db.session.query(GojekReport.transaction_id)
                    .filter(GojekReport.transaction_id.in_(batch))
                    .all()
                    if identifier
                )

            for batch in chunks(order_ids):
                existing_order_ids.update(
                    identifier
                    for (identifier,) in db.session.query(GojekReport.order_id)
                    .filter(GojekReport.order_id.in_(batch))
                    .all()
                    if identifier
                )

            return existing_transaction_ids, existing_order_ids

        def load_existing_gojek_amount_keys(rows):
            merchant_ids = set()
            transaction_dates = set()

            for row in rows:
                merchant_id = row.get('Merchant ID')
                if not merchant_id:
                    continue
                try:
                    transaction_date = datetime.strptime(row.get('Transaction Date', ''), '%m/%d/%Y').date()
                except ValueError:
                    continue
                merchant_ids.add(merchant_id)
                transaction_dates.add(transaction_date)

            if not merchant_ids or not transaction_dates:
                return set()

            existing_amount_keys = set()
            existing_reports = db.session.query(
                GojekReport.transaction_date,
                GojekReport.merchant_id,
                GojekReport.amount,
                GojekReport.nett_amount,
            ).filter(
                GojekReport.merchant_id.in_(merchant_ids),
                GojekReport.transaction_date.in_(transaction_dates),
            ).all()

            for transaction_date, merchant_id, amount, nett_amount in existing_reports:
                if amount is not None:
                    existing_amount_keys.add((transaction_date, merchant_id, float(amount)))
                if nett_amount is not None:
                    existing_amount_keys.add((transaction_date, merchant_id, float(nett_amount)))

            return existing_amount_keys

        def has_duplicate_gojek_identifier(transaction_id, order_id, existing_transaction_ids, existing_order_ids):
            if transaction_id and transaction_id in seen_transaction_ids:
                return True
            if order_id and order_id in seen_order_ids:
                return True
            return (
                (transaction_id and transaction_id in existing_transaction_ids)
                or (order_id and order_id in existing_order_ids)
            )

        def remember_gojek_identifiers(transaction_id, order_id):
            if transaction_id:
                seen_transaction_ids.add(transaction_id)
            if order_id:
                seen_order_ids.add(order_id)

        def has_duplicate_gojek_amount(transaction_date, merchant_id, amount, existing_amount_keys):
            if amount is None or not merchant_id:
                return False

            amount_key = (transaction_date, merchant_id, amount)
            return amount_key in seen_amount_keys or amount_key in existing_amount_keys

        def remember_gojek_amount(transaction_date, merchant_id, amount):
            if amount is not None and merchant_id:
                seen_amount_keys.add((transaction_date, merchant_id, amount))

        for file in files:
            file_contents = file.read().decode('utf-8')
            csv_file = StringIO(file_contents)
            reader = csv.DictReader(csv_file)
            rows = list(reader)

            transaction_ids = set()
            order_ids = set()
            for row in rows:
                transaction_id = clean_identifier(row.get('Transaction ID'))
                order_id = clean_identifier(row.get('Order ID'))
                if transaction_id:
                    transaction_ids.add(transaction_id)
                if order_id:
                    order_ids.add(order_id)

            existing_transaction_ids, existing_order_ids = load_existing_gojek_identifiers(
                transaction_ids,
                order_ids,
            )
            existing_amount_keys = load_existing_gojek_amount_keys(rows)
            
            reports = []
            for row in rows:
                merchant_name = row.get('Merchant name', '').strip()
                merchant_id = row.get('Merchant ID')
                if merchant_name and merchant_id:
                    store_id_map[merchant_name] = merchant_id

                outlet = None
                if merchant_id:
                    outlet = outlets_by_store_id.get(merchant_id)
                if not outlet and merchant_name:
                    outlet = outlets_by_name.get(merchant_name)
                if not outlet:
                    continue

                transaction_id = clean_identifier(row.get('Transaction ID'))
                order_id = clean_identifier(row.get('Order ID'))
                if has_duplicate_gojek_identifier(
                    transaction_id,
                    order_id,
                    existing_transaction_ids,
                    existing_order_ids,
                ):
                    skipped_reports += 1
                    continue

                try:
                    date_str = row.get('Transaction Date', '')
                    time_str = row.get('Transaction time', '')
                    
                    try:
                        transaction_date = datetime.strptime(date_str, '%m/%d/%Y').date()
                    except ValueError:
                        continue

                    try:
                        transaction_time = datetime.fromisoformat(time_str.replace('Z', '+00:00')).time()
                    except ValueError:
                        transaction_time = None

                    parsed_amount = safe_float_or_none(row.get('Amount', 0) or 0)
                    if has_duplicate_gojek_amount(transaction_date, merchant_id, parsed_amount, existing_amount_keys):
                        skipped_reports += 1
                        continue

                    report = {
                        'brand_name': outlet.brand,
                        'outlet_code': outlet.outlet_code,
                        'transaction_id': transaction_id or '',
                        'transaction_date': transaction_date,
                        'transaction_time': transaction_time,
                        'stan': row.get('Stan', ''),
                        'nett_amount': safe_float(row.get('Nett Amount')),
                        'amount': safe_float(row.get('Amount')),
                        'transaction_status': row.get('Transaction Status', ''),
                        'transaction_reference': clean_identifier(row.get('Transaction Reference')) or '',
                        'order_id': order_id or '',
                        'feature': row.get('Feature', ''),
                        'payment_type': row.get('Payment Type', ''),
                        'merchant_name': merchant_name,
                        'merchant_id': merchant_id,
                        'promo_type': row.get('Promo Type', ''),
                        'promo_name': row.get('Promo Name', ''),
                        'gopay_promo': safe_float(row.get('Gopay promo')),
                        'gofood_discount': safe_float(row.get('GoFood discount')),
                        'voucher_commission': safe_float(row.get('Voucher commission')),
                        'tax': safe_float(row.get('Tax')),
                        'witholding_tax': safe_float(row.get('Witholding tax')),
                        'currency': row.get('Currency', 'IDR')
                    }
                    reports.append(report)
                    remember_gojek_identifiers(transaction_id, order_id)
                    remember_gojek_amount(transaction_date, merchant_id, parsed_amount)
                    affected_outlets.add((outlet.outlet_code, transaction_date))
                    total_reports += 1
                except (ValueError, TypeError) as e:
                    print(f"Error processing row: {e}")
                    continue

            if reports:
                db.session.bulk_insert_mappings(GojekReport, reports)

        for outlet_id, date in affected_outlets:
            update_daily_total_for_outlet(outlet_id, date, 'gojek')

        db.session.commit()

        updated_count = update_store_ids_batch(store_id_map, 'gojek')
        
        return jsonify({
            'msg': 'Reports uploaded and consolidated successfully',
            'total_records': total_reports,
            'skipped_records': skipped_reports,
            'store_ids_updated': updated_count
        }), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@reports_bp.route('/upload/shopeepay', methods=['POST'])
def upload_report_shopeepay():
    files = request.files.getlist('file')
    if not files:
        files = [request.files.get('file')]
    
    if not files or not files[0]:
        return jsonify({'msg': 'No files uploaded'}), 400

    try:
        total_reports = 0
        skipped_reports = 0
        store_id_map = {}
        affected_outlets = set()

        for file in files:
            file_contents = file.read().decode('utf-8')
            csv_file = StringIO(file_contents)
            reader = csv.DictReader(csv_file)
            
            reports = []
            for row in reader:
                store_name = row.get('Merchant/Store Name', '').strip()
                entity_id = row.get('Entity ID')
                if store_name and entity_id:
                    store_id_map[store_name] = entity_id

                outlet = None
                if entity_id:
                    outlet = Outlet.query.filter_by(store_id_shopee=entity_id).first()
                if not outlet and store_name:
                    outlet = Outlet.query.filter_by(outlet_name_grab=store_name).first()
                if not outlet:
                    continue

                transaction_id = row.get('Transaction ID', '')
                if transaction_id:
                    existing_report = ShopeepayReport.query.filter_by(transaction_id=transaction_id).first()
                    if existing_report:
                        skipped_reports += 1
                        continue

                try:
                    create_time = datetime.strptime(row.get('Create Time', ''), '%Y-%m-%d %H:%M:%S')
                    update_time = None
                    if row.get('Update Time'):
                        update_time = datetime.strptime(row.get('Update Time', ''), '%Y-%m-%d %H:%M:%S')

                    def safe_float(value):
                        if not value:
                            return 0
                        try:
                            return float(str(value).replace(',', ''))
                        except (ValueError, TypeError):
                            return 0

                    report = ShopeepayReport(
                        brand_name=outlet.brand,
                        outlet_code=outlet.outlet_code,
                        merchant_host=row.get('Merchant Host', ''),
                        partner_merchant_id=row.get('Partner Merchant ID', ''),
                        merchant_store_name=store_name,
                        transaction_type=row.get('Transaction Type', ''),
                        merchant_scope=row.get('Merchant Scope', ''),
                        transaction_id=transaction_id,
                        reference_id=row.get('Reference ID', ''),
                        parent_id=row.get('Parent ID', ''),
                        external_reference_id=row.get('External Reference ID', ''),
                        issuer_identifier=row.get('Issuer Identifier', ''),
                        transaction_amount=safe_float(row.get('Transaction Amount')),
                        fee_mdr=safe_float(row.get('Fee (MDR)')),
                        settlement_amount=safe_float(row.get('Settlement Amount')),
                        terminal_id=row.get('Terminal ID', ''),
                        create_time=create_time,
                        update_time=update_time,
                        adjustment_reason=row.get('Adjustment Reason', ''),
                        entity_id=entity_id,
                        fee_cofunding=safe_float(row.get('Fee (Cofunding)')),
                        reward_amount=safe_float(row.get('Reward Amount')),
                        reward_type=row.get('Reward Type', ''),
                        promo_type=row.get('Promo Type', ''),
                        payment_method=row.get('Payment Method', ''),
                        currency_code=row.get('Currency Code', ''),
                        voucher_promotion_event_name=row.get('Voucher Promotion Event Name', ''),
                        payment_option=row.get('Payment Option', ''),
                        fee_withdrawal=safe_float(row.get('Fee (Withdrawal)')),
                        fee_handling=safe_float(row.get('Fee (Handling)'))
                    )
                    reports.append(report)
                    affected_outlets.add((outlet.outlet_code, create_time.date()))
                    total_reports += 1
                except (ValueError, TypeError) as e:
                    print(f"Error processing row: {e}")
                    continue

            if reports:
                db.session.bulk_save_objects(reports)

        for outlet_id, date in affected_outlets:
            update_daily_total_for_outlet(outlet_id, date, 'shopeepay')

        db.session.commit()

        updated_count = update_store_ids_batch(store_id_map, 'shopeepay')
        
        return jsonify({
            'msg': 'Reports uploaded and consolidated successfully',
            'total_records': total_reports,
            'skipped_records': skipped_reports,
            'store_ids_updated': updated_count
        }), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500
    
@reports_bp.route('/upload/tiktok', methods=['POST'])
def upload_report_tiktok():
    files = request.files.getlist('file')
    if not files:
        files = [request.files.get('file')]
    if not files or not files[0]:
        return jsonify({'msg': 'No files uploaded'}), 400

    try:
        total_reports = 0
        skipped_reports = 0
        debug_skipped = []

        for file in files:
            file_contents = file.read().decode('utf-8')
            csv_file = StringIO(file_contents)
            reader = csv.reader(csv_file)
            for _ in range(4):
                next(reader, None)  # Skip first 4 rows

            reports = []
            for idx, row in enumerate(reader):
                parsed = TiktokReport.parse_tiktok_row(row)

                if parsed:
                     # Duplicate check: adjust fields as needed for your business logic
                    exists = TiktokReport.query.filter_by(
                        store_name=parsed['store_name'],
                        outlet_order_id=parsed['outlet_order_id'],
                        order_time=parsed['order_time'],
                        settlement_time = parsed['settlement_time'],
                        gross_amount=parsed['gross_amount'],
                        net_amount=parsed['net_amount']
                    ).first()
                    if exists:
                        skipped_reports += 1
                        debug_skipped.append({
                            'row_number': idx + 2,
                            'reason': 'Duplicate entry',
                            'row': row
                        })
                        continue
                    report = TiktokReport(**parsed)
                    reports.append(report)
                    total_reports += 1
                else:
                    skipped_reports += 1
                    debug_skipped.append({
                        'row_number': idx + 2,
                        'reason': 'Parse failed or outlet not found',
                        'row': row
                    })

            if reports:
                db.session.bulk_save_objects(reports)
                db.session.commit()

        return jsonify({
            'msg': 'Tiktok reports uploaded successfully',
            'total_records': total_reports,
            'skipped_records': skipped_reports,
            'skipped_rows_debug': debug_skipped
        }), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500



@reports_bp.route('/upload/manual', methods=['POST'])
def upload_manual_entry():
    files = request.files.getlist('file')
    if not files:
        files = [request.files.get('file')]
    
    if not files or not files[0]:
        return jsonify({'msg': 'No files uploaded'}), 400

    try:
        total_entries = 0
        skipped_entries = 0
        
        for file in files:
            file_contents = file.read().decode('utf-8')
            csv_file = StringIO(file_contents)
            reader = csv.reader(csv_file)
            
            # Skip the first three rows (header and title)
            for _ in range(3):
                next(reader)
            
            entries = []
            for row in reader:
                try:
                    if len(row) < 7:  # Ensure we have all required columns
                        skipped_entries += 1
                        continue

                    # Skip empty or malformed date fields
                    date_str = row[0].strip()
                    if not date_str:
                        skipped_entries += 1
                        continue

                    try:
                        entry_date = parse_date(date_str, default_year=datetime.now().year)
                    except ValueError:
                        skipped_entries += 1
                        continue

                    # Parse amount (remove commas and convert to float)
                    amount_str = row[2].replace('.','').replace(',', '').strip()
                    if not amount_str:
                        skipped_entries += 1
                        continue

                    try:
                        amount = float(amount_str)
                    except ValueError:
                        skipped_entries += 1
                        continue

                    # Determine entry type and category
                    entry_type_str = row[3].strip().lower()
                    category_name = row[4].strip()
                    if entry_type_str == 'pengeluaran':
                        entry_type = 'expense'
                        category = ExpenseCategory.query.filter_by(name=category_name).first()
                    elif entry_type_str == 'penerimaan':
                        entry_type = 'income'
                        category = IncomeCategory.query.filter_by(name=category_name).first()
                    else:
                        skipped_entries += 1
                        continue

                    if not category:
                        skipped_entries += 1
                        continue

                    outlet_code = row[6].strip()
                    outlet = Outlet.query.filter_by(outlet_code=outlet_code).first()
                    start_date, end_date = resolve_upload_manual_entry_date_range(
                        outlet, entry_date
                    )
                    category_id = category.id if category else None

                    # Check for existing entry
                    existing_entry = ManualEntry.query.filter(
                        ManualEntry.start_date == start_date,
                        ManualEntry.end_date == end_date,
                        ManualEntry.outlet_code == outlet_code,
                        ManualEntry.category_id == category_id,
                        ManualEntry.amount == amount,
                        ManualEntry.description == row[5].strip()
                    ).first()

                    if existing_entry:
                        skipped_entries += 1
                        continue

                    # Create manual entry
                    entry = ManualEntry(
                        outlet_code=outlet_code,
                        brand_name='MP78',  # Assuming brand is always '78'
                        entry_type=entry_type,
                        amount=amount,
                        description=row[5].strip(),
                        start_date=start_date,
                        end_date=end_date,
                        category_id=category_id
                    )
                    entries.append(entry)
                    total_entries += 1


                except (ValueError, IndexError) as e:
                    print(f"Error processing row: {e}")
                    skipped_entries += 1
                    continue

            # Bulk save entries
            if entries:
                try:
                    db.session.bulk_save_objects(entries)
                    db.session.commit()
                except IntegrityError:
                    db.session.rollback()
                    # Handle entries one by one if bulk insert fails
                    for entry in entries:
                        try:
                            db.session.add(entry)
                            db.session.commit()
                        except IntegrityError:
                            db.session.rollback()
                            skipped_entries += 1
                        except Exception as e:
                            db.session.rollback()
                            print(f"Error saving entry: {str(e)}")

        return jsonify({
            'msg': 'Manual entries uploaded successfully',
            'total_records': total_entries,
            'skipped_records': skipped_entries
        }), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500
    

@reports_bp.route('/upload/grab', methods=['POST'])
def upload_report_grab():
    files = request.files.getlist('file')
    if not files:
        files = [request.files.get('file')]
    
    if not files or not files[0]:
        return jsonify({'msg': 'No files uploaded'}), 400

    try:
        total_reports = 0
        skipped_reports = 0
        store_id_map = {}
        affected_outlets = set()
        seen_transaction_ids = set()
        seen_long_order_ids = set()
        seen_short_order_ids = set()

        outlets = Outlet.query.all()
        outlets_by_store_id = {outlet.store_id_grab: outlet for outlet in outlets if outlet.store_id_grab}
        outlets_by_name = {outlet.outlet_name_grab: outlet for outlet in outlets if outlet.outlet_name_grab}

        def clean_identifier(value):
            if value is None:
                return None
            value = str(value).strip()
            return value or None

        def safe_float(value):
            if not value:
                return 0
            try:
                return float(str(value).replace(',', ''))
            except (ValueError, TypeError):
                return 0

        def chunks(values, size=1000):
            values = list(values)
            for index in range(0, len(values), size):
                yield values[index:index + size]

        def load_existing_grab_identifiers(transaction_ids, long_order_ids, short_order_ids):
            existing_transaction_ids = set()
            existing_long_order_ids = set()
            existing_short_order_ids = set()

            for batch in chunks(transaction_ids):
                existing_transaction_ids.update(
                    identifier
                    for (identifier,) in db.session.query(GrabFoodReport.id_transaksi)
                    .filter(GrabFoodReport.id_transaksi.in_(batch))
                    .all()
                    if identifier
                )

            for batch in chunks(long_order_ids):
                existing_long_order_ids.update(
                    identifier
                    for (identifier,) in db.session.query(GrabFoodReport.id_pesanan_panjang)
                    .filter(GrabFoodReport.id_pesanan_panjang.in_(batch))
                    .all()
                    if identifier
                )

            for batch in chunks(short_order_ids):
                existing_short_order_ids.update(
                    identifier
                    for (identifier,) in db.session.query(GrabFoodReport.id_pesanan_pendek)
                    .filter(GrabFoodReport.id_pesanan_pendek.in_(batch))
                    .all()
                    if identifier
                )

            return existing_transaction_ids, existing_long_order_ids, existing_short_order_ids

        def has_duplicate_grab_identifier(
            transaction_id,
            long_order_id,
            short_order_id,
            existing_transaction_ids,
            existing_long_order_ids,
            existing_short_order_ids,
        ):
            if transaction_id and transaction_id in seen_transaction_ids:
                return True
            if long_order_id and long_order_id in seen_long_order_ids:
                return True
            if short_order_id and short_order_id in seen_short_order_ids:
                return True
            return (
                (transaction_id and transaction_id in existing_transaction_ids)
                or (long_order_id and long_order_id in existing_long_order_ids)
                or (short_order_id and short_order_id in existing_short_order_ids)
            )

        def remember_grab_identifiers(transaction_id, long_order_id, short_order_id):
            if transaction_id:
                seen_transaction_ids.add(transaction_id)
            if long_order_id:
                seen_long_order_ids.add(long_order_id)
            if short_order_id:
                seen_short_order_ids.add(short_order_id)

        for file in files:
            file_contents = file.read().decode('utf-8')
            csv_file = StringIO(file_contents)
            reader = csv.DictReader(csv_file)
            rows = list(reader)

            transaction_ids = set()
            long_order_ids = set()
            short_order_ids = set()
            for row in rows:
                transaction_id = clean_identifier(row.get('ID transaksi'))
                long_order_id = clean_identifier(row.get('ID pesanan (panjang)'))
                short_order_id = clean_identifier(row.get('ID pesanan (pendek)'))
                if transaction_id:
                    transaction_ids.add(transaction_id)
                if long_order_id:
                    long_order_ids.add(long_order_id)
                if short_order_id:
                    short_order_ids.add(short_order_id)

            existing_transaction_ids, existing_long_order_ids, existing_short_order_ids = load_existing_grab_identifiers(
                transaction_ids,
                long_order_ids,
                short_order_ids,
            )
            
            reports = []
            for row in rows:
                store_name = row.get('Nama toko', '').strip()
                store_id = row.get('ID toko')
                if store_name and store_id:
                    store_id_map[store_name] = store_id

                outlet = None
                if store_id:
                    outlet = outlets_by_store_id.get(store_id)
                if not outlet and store_name:
                    outlet = outlets_by_name.get(store_name)
                if not outlet:
                    continue

                transaction_id = clean_identifier(row.get('ID transaksi'))
                long_order_id = clean_identifier(row.get('ID pesanan (panjang)'))
                short_order_id = clean_identifier(row.get('ID pesanan (pendek)'))
                if has_duplicate_grab_identifier(
                    transaction_id,
                    long_order_id,
                    short_order_id,
                    existing_transaction_ids,
                    existing_long_order_ids,
                    existing_short_order_ids,
                ):
                    skipped_reports += 1
                    continue

                try:
                    date_str = row.get('Tanggal dibuat', '')
                    date_made_str = row.get('Diperbarui Pada', '')
                    try:
                        tanggal_dibuat = datetime.strptime(date_str, '%d %b %Y %I:%M %p')
                        tanggal_diperbarui = datetime.strptime(date_made_str, '%d %b %Y %I:%M %p')

                    except ValueError:
                        tanggal_dibuat = datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S')
                        tanggal_diperbarui = datetime.strptime(date_made_str, '%Y-%m-%d %H:%M:%S')
                    amount = safe_float(row.get('Amount'))
                    total = safe_float(row.get('Total'))
                    if amount == 0 or total == 0:
                        skipped_reports += 1
                        continue

                    report = {
                        'brand_name': outlet.brand,
                        'outlet_code': outlet.outlet_code,
                        'nama_toko': store_name,
                        'id_toko': store_id,
                        'tanggal_dibuat': tanggal_dibuat,
                        'diperbarui_pada': tanggal_diperbarui,
                        'jenis': row.get('Jenis', ''),
                        'kategori': row.get('Kategori', ''),
                        'subkategori': row.get('Subkategori', ''),
                        'status': row.get('Status', ''),
                        'id_transaksi': transaction_id,
                        'id_pesanan_panjang': long_order_id,
                        'id_pesanan_pendek': short_order_id,
                        'komisi_grabkitchen': safe_float(row.get('Komisi GrabKitchen')),
                        'total': total,
                        'amount': amount,
                        'penjualan_bersih': safe_float(row.get('Penjualan bersih')),
                    }
                    reports.append(report)
                    remember_grab_identifiers(transaction_id, long_order_id, short_order_id)
                    affected_outlets.add((outlet.outlet_code, tanggal_diperbarui.date()))
                    total_reports += 1
                except (ValueError, TypeError) as e:
                    print(f"Error processing row: {e}")
                    continue

            if reports:
                db.session.bulk_insert_mappings(GrabFoodReport, reports)

        for outlet_id, date in affected_outlets:
            update_daily_total_for_outlet(outlet_id, date, 'grab')

        db.session.commit()

        updated_count = update_store_ids_batch(store_id_map, 'grab')
        
        return jsonify({
            'msg': 'Reports uploaded and consolidated successfully',
            'total_records': total_reports,
            'skipped_records': skipped_reports,
            'store_ids_updated': updated_count
        }), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@reports_bp.route('/upload/shopee', methods=['POST'])
def upload_report_shopee():
    files = request.files.getlist('file')
    if not files:
        files = [request.files.get('file')]
    
    if not files or not files[0]:
        return jsonify({'msg': 'No files uploaded'}), 400

    try:
        total_reports = 0
        skipped_reports = 0
        store_id_map = {}
        affected_outlets = set()

        for file in files:
            file_contents = file.read().decode('utf-8')
            csv_file = StringIO(file_contents)
            reader = csv.DictReader(csv_file)
            
            reader.fieldnames = [field.strip() for field in reader.fieldnames]
            
            reports = []
            for row in reader:
                store_name = row.get('Store Name', '').strip()
                store_id = row.get('Store ID')
                if store_name and store_id:
                    store_id_map[store_name] = store_id

                outlet = None
                if store_id:
                    outlet = Outlet.query.filter_by(store_id_shopee=store_id).first()
                if not outlet and store_name:
                    outlet = Outlet.query.filter_by(outlet_name_grab=store_name).first()
                if not outlet:
                    continue

                order_id = row.get('Order ID', '')
                if order_id:
                    existing_report = ShopeeReport.query.filter_by(order_id=order_id).first()
                    if existing_report:
                        skipped_reports += 1
                        continue

                try:
                    create_time = datetime.strptime(row.get('Order Create Time', ''), '%d/%m/%Y %H:%M:%S')
                    complete_time = None
                    if row.get('Order Complete/Cancel Time'):
                        complete_time = datetime.strptime(row.get('Order Complete/Cancel Time', ''), '%d/%m/%Y %H:%M:%S')

                    report = ShopeeReport(
                        brand_name=outlet.brand,
                        outlet_code=outlet.outlet_code,
                        transaction_type=row.get('Transaction Type', ''),
                        order_id=row.get('Order ID', ''),
                        order_pick_up_id=row.get('Order Pick up ID', ''),
                        store_id=store_id,
                        store_name=store_name,
                        order_create_time=create_time,
                        order_complete_cancel_time=complete_time,
                        order_amount=float(row.get('Order Amount', 0) or 0),
                        merchant_service_charge=float(row.get('Merchant Service Charge', 0) or 0),
                        pb1=float(row.get('PB1', 0) or 0),
                        merchant_surcharge_fee=float(row.get('Merchant Surcharge Fee', 0) or 0),
                        merchant_shipping_fee_voucher_subsidy=float(row.get('Merchant Shipping Fee Voucher Subsidy', 0) or 0),
                        food_direct_discount=float(row.get('Food Direct Discount', 0) or 0),
                        merchant_food_voucher_subsidy=float(row.get('Merchant Food Voucher Subsidy', 0) or 0),
                        subtotal=float(row.get('Subtotal', 0) or 0),
                        total=float(row.get('Total', 0) or 0),
                        commission=float(row.get('Commission', 0) or 0),
                        net_income=float(row.get('Net Income', 0) or 0),
                        order_status=row.get('Order Status', ''),
                        order_type=row.get('Order Type', '')
                    )
                    reports.append(report)
                    affected_outlets.add((outlet.outlet_code, create_time.date()))
                    total_reports += 1
                except (ValueError, TypeError) as e:
                    print(f"Error processing row: {e}")
                    continue

            if reports:
                db.session.bulk_save_objects(reports)

        for outlet_id, date in affected_outlets:
            update_daily_total_for_outlet(outlet_id, date, 'shopee')

        db.session.commit()

        updated_count = update_store_ids_batch(store_id_map, 'shopee')
        
        return jsonify({
            'msg': 'Reports uploaded and consolidated successfully',
            'total_records': total_reports,
            'skipped_records': skipped_reports,
            'store_ids_updated': updated_count
        }), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@reports_bp.route('/upload/cash', methods=['POST'])
def upload_cash_report():
    file = request.files.get('file')
    outlet_code = request.form.get('outlet_code')
    brand_name = request.form.get('brand_name')
    
    if not file:
        return jsonify({'msg': 'No file uploaded'}), 400
    if not outlet_code:
        return jsonify({'msg': 'Outlet code is required'}), 400

    pukis_mapping = {
        'Pukis Produksi': ('produksi', 'jumbo'),
        'Pukis Terjual Total Jumbo': ('terjual', 'jumbo'),
        'Pukis Retur': ('retur', 'jumbo'),
        'Pukis Free': ('free', 'jumbo'),
        'Pukis Produksi KLASIK': ('produksi', 'klasik'),
        'Pukis Terjual Total KLASIK': ('terjual', 'klasik'),
        'Pukis Retur KLASIK': ('retur', 'klasik'),
        'Pukis Free KLASIK': ('free', 'klasik'),
    }

    try:
        file_contents = file.read().decode('utf-8')
        csv_file = StringIO(file_contents)
        reader = csv.reader(csv_file)
        reports = []
        pukis_reports = []
        skipped_duplicates = 0
        skipped_pukis_duplicates = 0
        debug_skipped = []
        header = next(reader, None)  # Skip header row
        for idx, row in enumerate(reader):
            if len(row) < 5:
                debug_skipped.append({
                    'row_number': idx + 2,
                    'reason': 'Not enough columns',
                    'row': row
                })
                continue
            date_str = row[0].strip()
            keterangan1 = row[2].strip()
            details = row[3].strip()
            total_str = row[4].strip().replace('.', '').replace(',', '')
            if not date_str or not keterangan1 or not total_str:
                debug_skipped.append({
                    'row_number': idx + 2,
                    'reason': 'Missing required fields',
                    'row': row
                })
                continue
            # Handle Pukis rows
            if keterangan1 in pukis_mapping:
                try:
                    tanggal = datetime.strptime(date_str, '%d %b %Y')
                    amount = float(total_str)
                    pukis_inventory_type, pukis_product_type = pukis_mapping[keterangan1]
                    # Duplicate check for Pukis
                    existing_pukis = Pukis.query.filter_by(
                        tanggal=tanggal,
                        outlet_code=outlet_code,
                        brand_name=brand_name,
                        pukis_inventory_type=pukis_inventory_type,
                        pukis_product_type=pukis_product_type
                    ).first()
                    if existing_pukis:
                        skipped_pukis_duplicates += 1
                        debug_skipped.append({
                            'row_number': idx + 2,
                            'reason': 'Duplicate Pukis entry',
                            'row': row
                        })
                        continue
                    pukis_report = Pukis(
                        tanggal=tanggal,
                        outlet_code=outlet_code,
                        brand_name=brand_name,
                        pukis_inventory_type=pukis_inventory_type,
                        pukis_product_type=pukis_product_type,
                        amount=amount
                    )
                    pukis_reports.append(pukis_report)
                except Exception as e:
                    debug_skipped.append({
                        'row_number': idx + 2,
                        'reason': f'Pukis parse error: {str(e)}',
                        'row': row
                    })
                continue  # Don't process as CashReport
            # ...existing code for cash report rows...
            type_str = keterangan1.lower()
            if type_str not in ['penerimaan', 'pengeluaran']:
                debug_skipped.append({
                    'row_number': idx + 2,
                    'reason': f'Unknown type: {type_str}',
                    'row': row
                })
                continue
            try:
                date = datetime.strptime(date_str, '%d %b %Y')
                total = float(total_str)
            except Exception as e:
                debug_skipped.append({
                    'row_number': idx + 2,
                    'reason': f'Parse error: {str(e)}',
                    'row': row
                })
                continue
            existing_record = CashReport.query.filter(
                CashReport.tanggal == date,
                CashReport.total == total,
                CashReport.outlet_code == outlet_code
            ).first()
            if existing_record:
                skipped_duplicates += 1
                debug_skipped.append({
                    'row_number': idx + 2,
                    'reason': 'Duplicate entry',
                    'row': row
                })
                continue
            type_mapping = {
                'penerimaan': 'income',
                'pengeluaran': 'expense'
            }
            report = CashReport(
                tanggal=date,
                outlet_code=outlet_code,
                brand_name=brand_name,
                type=type_mapping[type_str],
                details=details,
                total=total
            )
            reports.append(report)
        if reports:
            db.session.bulk_save_objects(reports)
        if pukis_reports:
            db.session.bulk_save_objects(pukis_reports)
        if reports or pukis_reports:
            db.session.commit()
        return jsonify({
            'msg': 'Reports uploaded successfully',
            'cash_count': len(reports),
            'pukis_count': len(pukis_reports),
            'skipped_duplicates': skipped_duplicates,
            'skipped_pukis_duplicates': skipped_pukis_duplicates,
            'skipped_rows_debug': debug_skipped
        }), 201
    except IntegrityError:
        db.session.rollback()
        return jsonify({'msg': 'Duplicate entry error'}), 500
    except Exception as e:
        db.session.rollback()
        return jsonify({'msg': str(e)}), 500

@reports_bp.route('/upload/pkb', methods=['POST'])
def upload_report_pkb():
    """
    Handles the upload of a PKB bank mutation report.
    This function delegates the parsing of the report to the BankMutation model.
    """
    files = request.files.getlist('file')
    rekening_number = request.form.get('rekening_number')

    if not files:
        files = [request.files.get('file')]
    if not files or not files[0]:
        return jsonify({'msg': 'No files uploaded'}), 400
    if not rekening_number:
        return jsonify({'msg': 'Rekening number is required'}), 400

    try:
        total_mutations = 0
        skipped_mutations = 0

        for file in files:
            if not file.filename.endswith('.csv'):
                continue

            # Decode the file content as UTF-8 and handle potential BOM
            report_content = file.read().decode('utf-8-sig')

            # Delegate parsing to the BankMutation model
            parsed_data = BankMutation.parse_pkb_report(report_content)
            # print(f"DEBUG Parsed Data: {parsed_data}")  # Debug log

            # Process the parsed transaction data
            for transaction_data in parsed_data.get('transactions', []):
                transaction_data['rekening_number'] = rekening_number

                # Check for duplicates
                exists = BankMutation.query.filter_by(
                    tanggal=transaction_data.get('tanggal'),
                    transaction_amount=transaction_data.get('transaction_amount'),
                    platform_code=transaction_data.get('platform_code'),
                    transaksi=transaction_data.get('transaksi')
                ).first()
                
                if not exists:
                    new_mutation = BankMutation(**transaction_data)
                    db.session.add(new_mutation)
                    total_mutations += 1
                else:
                    skipped_mutations += 1

        db.session.commit()

        return jsonify({
            'msg': 'PKB report uploaded and processed successfully',
            'total_records': total_mutations,
            'skipped_records': skipped_mutations
        }), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error processing file: {str(e)}'}), 500


@reports_bp.route('/commission-totals', methods=['GET'])
def get_commission_totals():
    start_date_param = request.args.get('start_date')
    end_date_param = request.args.get('end_date')
    outlet_code = request.args.get('outlet_code')
    brand_name = request.args.get('brand_name')

    try:
        # Initialize query for GrabFood reports
        grab_query = GrabFoodReport.query

        

        # Apply filters based on outlet_code and brand_name
        if outlet_code and outlet_code.upper() != "ALL":
            grab_query = grab_query.filter(GrabFoodReport.outlet_code == outlet_code)
        
        if brand_name and brand_name != "ALL":
             grab_query = grab_query.filter(GrabFoodReport.brand_name.in_(["MP78", "MP78 Express"]))
        if brand_name and brand_name not in ["MP78", "MP78 Express"]:
            return jsonify({'error': 'Commission calculation is only available for MP78 brands'}), 400
        else:
            # Only calculate commission for MP78 brands if no specific brand is selected
            grab_query = grab_query.filter(GrabFoodReport.brand_name.in_(["MP78", "MP78 Express"]))

        # Apply date filters
        if start_date_param and end_date_param:
            start_date = datetime.strptime(start_date_param, '%Y-%m-%d')
            end_date = datetime.strptime(end_date_param, '%Y-%m-%d')
            end_date_inclusive = datetime(end_date.year, end_date.month, end_date.day, 23, 59, 59)
            
            grab_query = grab_query.filter(
                GrabFoodReport.tanggal_dibuat >= start_date,
                GrabFoodReport.tanggal_dibuat <= end_date_inclusive,
                GrabFoodReport.jenis == "GrabFood"
            )

        # Calculate totals and commission
        grab_total = sum(float(report.total or 0) for report in grab_query.all())
        management_commission = round(grab_total * 1/74, 2)  # 1% commission
        partner_commission = round(grab_total * 1/74, 2)    # 1% commission

        response = {
            'period': {
                'start_date': start_date_param,
                'end_date': end_date_param
            },
            'outlet_code': outlet_code,
            'brand_name': brand_name,
            'totals': {
                'grab_food_total': round(grab_total, 2),
                'management_commission': management_commission,
                'partner_commission': partner_commission,
                'total_commission': management_commission + partner_commission
            }
        }
        
        return jsonify(response), 200

    except ValueError as e:
        return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400
    except Exception as e:
        print(f"Error calculating commission totals: {str(e)}")
        return jsonify({'error': str(e)}), 500


@reports_bp.route('/totals', methods=['GET'])
def get_reports_totals():
    start_date_param = request.args.get('start_date')
    end_date_param = request.args.get('end_date')
    outlet_code = request.args.get('outlet_code')
    brand_name = request.args.get('brand_name')

    try:
        from app.models.qpon_reports import QponReport

        # Initialize queries
        gojek_query = GojekReport.query
        grab_query = GrabFoodReport.query
        shopee_query = ShopeeReport.query
        shopeepay_query = ShopeepayReport.query
        tiktok_query = TiktokReport.query
        qpon_query = QponReport.query
        webshop_query = WebshopReport.query
        cash_query = CashReport.query
        manual_entries_query = ManualEntry.query
        mpr_totals = None
        # Return error if outlet_code is not provided
        if not outlet_code:
            return jsonify({'error': 'outlet_code is required'}), 400
         # Apply outlet_code filter if provided and not "ALL"
        if outlet_code.upper() != "ALL":
            gojek_query = gojek_query.filter(GojekReport.outlet_code == outlet_code)
            grab_query = grab_query.filter(GrabFoodReport.outlet_code == outlet_code)
            shopee_query = shopee_query.filter(ShopeeReport.outlet_code == outlet_code)
            shopeepay_query = shopeepay_query.filter(ShopeepayReport.outlet_code == outlet_code)
            tiktok_query = tiktok_query.filter(TiktokReport.outlet_code == outlet_code)
            qpon_query = qpon_query.filter(QponReport.outlet_code == outlet_code)
            webshop_query = WebshopReport.query.filter(WebshopReport.outlet_code == outlet_code)
            cash_query = cash_query.filter(CashReport.outlet_code == outlet_code)
            manual_entries_query = manual_entries_query.filter(ManualEntry.outlet_code == outlet_code)

        if outlet_code.upper() == "ALL" and brand_name != "ALL":
            gojek_query = gojek_query.filter(GojekReport.brand_name == brand_name)
            grab_query = grab_query.filter(GrabFoodReport.brand_name == brand_name)
            shopee_query = shopee_query.filter(ShopeeReport.brand_name == brand_name)
            cash_query = cash_query.filter(CashReport.brand_name == brand_name)
            shopeepay_query = shopeepay_query.filter(ShopeepayReport.brand_name == brand_name)
            tiktok_query = tiktok_query.filter(TiktokReport.brand_name == brand_name)
            qpon_query = qpon_query.filter(QponReport.brand_name == brand_name)
            webshop_query = WebshopReport.query.filter(WebshopReport.brand_name == brand_name)
            manual_entries_query = manual_entries_query.filter(ManualEntry.brand_name == brand_name)

        # Default cash queries (used when date filter is not provided)
        cash_income_query = cash_query.filter(CashReport.type == 'income')
        cash_expense_query = cash_query.filter(CashReport.type == 'expense')

        # Apply date filters if provided
        if start_date_param and end_date_param:
            start_date = datetime.strptime(start_date_param, '%Y-%m-%d')
            end_date = datetime.strptime(end_date_param, '%Y-%m-%d')
            
            # Add one day to end_date to include the entire end date
            end_date_inclusive = datetime(end_date.year, end_date.month, end_date.day, 23, 59, 59)
            
            gojek_query = gojek_query.filter(
                GojekReport.transaction_date >= start_date,
                GojekReport.transaction_date <= end_date_inclusive
            )
            grab_query = grab_query.filter(
                GrabFoodReport.tanggal_dibuat >= start_date,
                GrabFoodReport.tanggal_dibuat <= end_date_inclusive
            )
            shopee_query = shopee_query.filter(
                ShopeeReport.order_create_time >= start_date,
                ShopeeReport.order_create_time <= end_date_inclusive
            )
            shopeepay_query = shopeepay_query.filter(
                ShopeepayReport.create_time >= start_date,
                ShopeepayReport.create_time <= end_date_inclusive
            )
            tiktok_query = tiktok_query.filter(
                TiktokReport.order_time >= start_date,
                TiktokReport.order_time <= end_date_inclusive
            )
            qpon_query = qpon_query.filter(
                QponReport.bill_created_at >= start_date,
                QponReport.bill_created_at <= end_date_inclusive
            )
            webshop_query = webshop_query.filter(
                WebshopReport.created_at >= start_date,
                WebshopReport.created_at <= end_date_inclusive
            )
            # Use SQLAlchemy filtering for cash reports instead of post-filtering
            cash_income_query = cash_query.filter(
                CashReport.type == 'income',
                CashReport.tanggal >= start_date,
                CashReport.tanggal <= end_date_inclusive
            )
            
            cash_expense_query = cash_query.filter(
                CashReport.type == 'expense',
                CashReport.tanggal >= start_date,
                CashReport.tanggal <= end_date_inclusive
            )
            
            manual_entries_query = manual_entries_query.filter(
                ManualEntry.start_date >= start_date,
                ManualEntry.end_date <= end_date_inclusive
            )
        else:
            start_date = None
            end_date_inclusive = None

        if outlet_code.upper() != "ALL":
            outlet, mapping = get_mpr_mapping_for_outlet(outlet_code)
            if outlet and outlet.brand in ("MP78", "MPR") and mapping:
                mpr_outlet_code = mapping.mpr_outlet_code
                mpr_totals = calculate_mpr_totals(
                    mpr_outlet_code,
                    start_date,
                    end_date_inclusive,
                )

        # Calculate totals for each platform
        gojek_total = sum(float(report.nett_amount or 0) for report in gojek_query.all())
        grab_total = sum(float(report.total or 0) for report in grab_query.all())
        shopee_reports = shopee_query.all()
        shopee_total = sum(float(report.net_income or 0) for report in shopee_reports if report.order_status != "Cancelled")
        shopeepay_total = sum(
            float(report.settlement_amount or 0) 
            for report in shopeepay_query.all() 
            if report.transaction_type != "Withdrawal"
        )
        tiktok_total = sum(float(report.net_amount or 0) for report in tiktok_query.all())   
        qpon_total = sum(
            float((getattr(report, 'net_amount', None) if getattr(report, 'net_amount', None) is not None else getattr(report, 'nett_amount', 0)) or 0)
            for report in qpon_query.all()
        )
        webshop_total = sum(float(report.nett_value or 0) for report in webshop_query.all())
        
        # Calculate cash totals using the filtered queries
        cash_income = sum(float(report.total or 0) for report in cash_income_query.all())
        cash_expense = sum(float(report.total or 0) for report in cash_expense_query.all())
        cash_net = cash_income - cash_expense

        # Calculate manual entries totals (expenses)
        manual_entries_total = sum(float(entry.amount or 0) for entry in manual_entries_query.all())

        # Calculate running total
        running_total = gojek_total + grab_total + shopee_total + shopeepay_total + tiktok_total + qpon_total + webshop_total + cash_net - manual_entries_total
        response = {
            'outlet_code': outlet_code,
            'brand_name': brand_name,
            'period': {
                'start_date': start_date_param,
                'end_date': end_date_param
            },
            'totals': {
                'gojek': round(gojek_total, 2),
                'grab': round(grab_total, 2),
                'shopee': round(shopee_total, 2),
                'shopeepay': round(shopeepay_total, 2),
                'tiktok': round(tiktok_total, 2),
                'qpon': round(qpon_total, 2),
                'webshop': round(webshop_total, 2),
                'cash': {
                    'income': round(cash_income, 2),
                    'expense': round(cash_expense, 2),
                    'net': round(cash_net, 2)
                },
                'manual_entries': round(manual_entries_total, 2),
                'running_total': round(running_total, 2),
                # 'mp78_commission': mp78_commission,  # Added MP78 commission
                # 'mp78_grab_total': round(mp78_grab_total, 2)  # Added for reference
            }
        }
        if mpr_totals:
            response['totals']['mpr_totals'] = mpr_totals
        return jsonify(response), 200

    except ValueError as e:
        return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400
    except Exception as e:
        print(f"Error calculating totals: {str(e)}")
        return jsonify({'error': str(e)}), 500


@reports_bp.route('/failed-cancelled-transfers', methods=['GET'])
def get_failed_cancelled_transfers():
    start_date_param = request.args.get('start_date')
    end_date_param = request.args.get('end_date')
    outlet_code = request.args.get('outlet_code')
    brand_name = request.args.get('brand_name')
    platform = (request.args.get('platform') or '').lower()
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    successful_statuses = ('Ditransfer', 'Transferred', 'Completed', 'Selesai')

    if not start_date_param or not end_date_param:
        return jsonify({'error': 'start_date and end_date are required'}), 400

    if not platform:
        return jsonify({'error': 'platform is required'}), 400

    if platform != 'grab':
        return jsonify({'error': f'Platform "{platform}" is not supported yet'}), 400

    if page < 1:
        return jsonify({'error': 'page must be greater than or equal to 1'}), 400

    if per_page < 1 or per_page > 100:
        return jsonify({'error': 'per_page must be between 1 and 100'}), 400

    try:
        start_date = datetime.strptime(start_date_param, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_param, '%Y-%m-%d')
        end_date_inclusive = datetime(end_date.year, end_date.month, end_date.day, 23, 59, 59)

        grab_query = GrabFoodReport.query.filter(
            GrabFoodReport.tanggal_dibuat >= start_date,
            GrabFoodReport.tanggal_dibuat <= end_date_inclusive,
            db.or_(
                GrabFoodReport.status.is_(None),
                ~GrabFoodReport.status.in_(successful_statuses)
            )
        )

        if outlet_code and outlet_code.upper() != "ALL":
            grab_query = grab_query.filter(GrabFoodReport.outlet_code == outlet_code)

        if brand_name and brand_name.upper() != "ALL":
            grab_query = grab_query.filter(GrabFoodReport.brand_name == brand_name)

        total_records = grab_query.count()
        total_pages = (total_records + per_page - 1) // per_page
        amount_total, net_total = grab_query.with_entities(
            func.coalesce(func.sum(GrabFoodReport.amount), 0),
            func.coalesce(func.sum(GrabFoodReport.total), 0)
        ).first()

        reports = grab_query.order_by(GrabFoodReport.tanggal_dibuat.asc()).offset(
            (page - 1) * per_page
        ).limit(per_page).all()
        transactions = [
            {
                'id': report.id,
                'brand_name': report.brand_name,
                'outlet_code': report.outlet_code,
                'nama_toko': report.nama_toko,
                'tanggal_dibuat': report.tanggal_dibuat.isoformat() if report.tanggal_dibuat else None,
                'tanggal_transfer': report.tanggal_transfer.isoformat() if report.tanggal_transfer else None,
                'status': report.status,
                'amount': float(report.amount or 0),
                'total': float(report.total or 0),
            }
            for report in reports
        ]

        return jsonify({
            'platform': platform,
            'period': {
                'start_date': start_date_param,
                'end_date': end_date_param
            },
            'outlet_code': outlet_code,
            'brand_name': brand_name,
            'excluded_statuses': list(successful_statuses),
            'count': len(transactions),
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total_records': total_records,
                'total_pages': total_pages,
                'has_next': page < total_pages,
                'has_prev': page > 1
            },
            'totals': {
                'amount': round(float(amount_total or 0), 2),
                'total': round(float(net_total or 0), 2)
            },
            'transactions': transactions
        }), 200

    except ValueError:
        return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400
    except Exception as e:
        print(f"Error fetching failed/cancelled transfers: {str(e)}")
        return jsonify({'error': str(e)}), 500

@reports_bp.route('/top-outlets', methods=['GET'])
def get_top_outlets():
    start_date_param = request.args.get('start_date')
    end_date_param = request.args.get('end_date')
    brand_name = request.args.get('brand_name')
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 20))

    if not start_date_param or not end_date_param or not brand_name:
        return jsonify({'error': 'start_date, end_date, and brand_name are required'}), 400

    try:
        start_date = datetime.strptime(start_date_param, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_param, '%Y-%m-%d')
        end_date_inclusive = datetime(end_date.year, end_date.month, end_date.day, 23, 59, 59)

        # Fetch all relevant data by brand and date
        gojek_data = GojekReport.query.filter(
            GojekReport.brand_name == brand_name,
            GojekReport.transaction_date >= start_date,
            GojekReport.transaction_date <= end_date_inclusive
        ).all()

        grab_data = GrabFoodReport.query.filter(
            GrabFoodReport.brand_name == brand_name,
            GrabFoodReport.tanggal_dibuat >= start_date,
            GrabFoodReport.tanggal_dibuat <= end_date_inclusive
        ).all()

        shopee_data = ShopeeReport.query.filter(
            ShopeeReport.brand_name == brand_name,
            ShopeeReport.order_create_time >= start_date,
            ShopeeReport.order_create_time <= end_date_inclusive
        ).all()

        shopeepay_data = ShopeepayReport.query.filter(
            ShopeepayReport.brand_name == brand_name,
            ShopeepayReport.create_time >= start_date,
            ShopeepayReport.create_time <= end_date_inclusive
        ).all()

        cash_data = CashReport.query.filter(
            CashReport.brand_name == brand_name,
            CashReport.tanggal >= start_date,
            CashReport.tanggal <= end_date_inclusive
        ).all()

        manual_entries_data = ManualEntry.query.filter(
            ManualEntry.brand_name == brand_name,
            ManualEntry.start_date >= start_date,
            ManualEntry.end_date <= end_date_inclusive
        ).all()

        outlet_totals = {}

        # Helper to accumulate totals
        def add_total(outlet, amount):
            outlet_totals[outlet] = outlet_totals.get(outlet, 0) + amount

        for report in gojek_data:
            add_total(report.outlet_code, float(report.nett_amount or 0))

        for report in grab_data:
            add_total(report.outlet_code, float(report.total or 0))

        for report in shopee_data:
            if report.order_status != "Cancelled":
                add_total(report.outlet_code, float(report.net_income or 0))

        for report in shopeepay_data:
            if report.transaction_type != "Withdrawal":
                add_total(report.outlet_code, float(report.settlement_amount or 0))

        # For cash: income - expense
        cash_summary = {}
        for report in cash_data:
            code = report.outlet_code
            amount = float(report.total or 0)
            if code not in cash_summary:
                cash_summary[code] = {'income': 0, 'expense': 0}
            if report.type == 'income':
                cash_summary[code]['income'] += amount
            elif report.type == 'expense':
                cash_summary[code]['expense'] += amount

        for code, summary in cash_summary.items():
            net_cash = summary['income'] - summary['expense']
            add_total(code, net_cash)

        # Subtract manual entry expenses
        for entry in manual_entries_data:
            add_total(entry.outlet_code, -float(entry.amount or 0))

        # Sort outlets by running total descending
        top_outlets = sorted(outlet_totals.items(), key=lambda x: x[1], reverse=True)
        # Filter to only valid outlets with matching brand and non-null name BEFORE pagination
        valid_outlets = []
        for outlet, total in top_outlets:
            outlet_obj = Outlet.query.filter_by(outlet_code=outlet, brand=brand_name).first()
            if outlet_obj and outlet_obj.brand == brand_name and outlet_obj.outlet_name_gojek:
                valid_outlets.append({
                    'outlet_code': outlet,
                    'outlet_brand': outlet_obj.brand,
                    'outlet_name': outlet_obj.outlet_name_gojek,
                    'running_total': round(total, 2)
                })

        total_records = len(valid_outlets)
        total_pages = (total_records + per_page - 1) // per_page
        start = (page - 1) * per_page
        end = start + per_page
        paginated_outlets = valid_outlets[start:end]

        response = {
            'top_outlets': paginated_outlets,
            'pagination': {
                'current_page': page,
                'per_page': per_page,
                'total_pages': total_pages,
                'total_records': total_records
            }
        }

        return jsonify(response), 200

    except ValueError:
        return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400
    except Exception as e:
        print(f"Error in /top-outlets: {str(e)}")
        return jsonify({'error': str(e)}), 500

@reports_bp.route('/top-outlets/pdf', methods=['GET'])
def get_top_outlets_pdf():
    start_date_param = request.args.get('start_date')
    end_date_param = request.args.get('end_date')
    brand_name = request.args.get('brand_name')

    if not start_date_param or not end_date_param or not brand_name:
        return jsonify({'error': 'start_date, end_date, and brand_name are required'}), 400

    start_date = datetime.strptime(start_date_param, '%Y-%m-%d')
    end_date = datetime.strptime(end_date_param, '%Y-%m-%d')
    end_date_inclusive = datetime(end_date.year, end_date.month, end_date.day, 23, 59, 59)

    # Fetch all relevant data by brand and date
    gojek_data = GojekReport.query.filter(
        GojekReport.brand_name == brand_name,
        GojekReport.transaction_date >= start_date,
        GojekReport.transaction_date <= end_date_inclusive
    ).all()

    grab_data = GrabFoodReport.query.filter(
        GrabFoodReport.brand_name == brand_name,
        GrabFoodReport.tanggal_dibuat >= start_date,
        GrabFoodReport.tanggal_dibuat <= end_date_inclusive
    ).all()

    shopee_data = ShopeeReport.query.filter(
        ShopeeReport.brand_name == brand_name,
        ShopeeReport.order_create_time >= start_date,
        ShopeeReport.order_create_time <= end_date_inclusive
    ).all()

    shopeepay_data = ShopeepayReport.query.filter(
        ShopeepayReport.brand_name == brand_name,
        ShopeepayReport.create_time >= start_date,
        ShopeepayReport.create_time <= end_date_inclusive
    ).all()

    cash_data = CashReport.query.filter(
        CashReport.brand_name == brand_name,
        CashReport.tanggal >= start_date,
        CashReport.tanggal <= end_date_inclusive
    ).all()

    manual_entries_data = ManualEntry.query.filter(
        ManualEntry.brand_name == brand_name,
        ManualEntry.start_date >= start_date,
        ManualEntry.end_date <= end_date_inclusive
    ).all()

    outlet_totals = {}

    def add_total(outlet, amount):
        outlet_totals[outlet] = outlet_totals.get(outlet, 0) + amount

    for report in gojek_data:
        add_total(report.outlet_code, float(report.nett_amount or 0))

    for report in grab_data:
        add_total(report.outlet_code, float(report.total or 0))

    for report in shopee_data:
        if report.order_status != "Cancelled":
            add_total(report.outlet_code, float(report.net_income or 0))

    for report in shopeepay_data:
        if report.transaction_type != "Withdrawal":
            add_total(report.outlet_code, float(report.settlement_amount or 0))

    # For cash: income - expense
    cash_summary = {}
    for report in cash_data:
        code = report.outlet_code
        amount = float(report.total or 0)
        if code not in cash_summary:
            cash_summary[code] = {'income': 0, 'expense': 0}
        if report.type == 'income':
            cash_summary[code]['income'] += amount
        elif report.type == 'expense':
            cash_summary[code]['expense'] += amount

    for code, summary in cash_summary.items():
        net_cash = summary['income'] - summary['expense']
        add_total(code, net_cash)

    # Subtract manual entry expenses
    for entry in manual_entries_data:
        add_total(entry.outlet_code, -float(entry.amount or 0))

    # Sort outlets by running total descending
    top_outlets = sorted(outlet_totals.items(), key=lambda x: x[1], reverse=True)
    # Filter to only valid outlets with matching brand and non-null name BEFORE pagination
    valid_outlets = []
    for outlet, total in top_outlets:
        outlet_obj = Outlet.query.filter_by(outlet_code=outlet, brand=brand_name, status='Active').first()
        if outlet_obj and outlet_obj.brand == brand_name and outlet_obj.outlet_name_gojek:
            valid_outlets.append({
                'outlet_code': outlet,
                'outlet_brand': outlet_obj.brand,
                'outlet_name': outlet_obj.outlet_name_gojek,
                'running_total': round(total, 2)
            })

    # Prepare table data: rank, outlet name, net income
    TABLE_HEADER = ("Rank", "Outlet Name", "Net Income (Rp)")
    table_rows = [TABLE_HEADER]
    for idx, outlet in enumerate(valid_outlets, 1):
        table_rows.append((str(idx), outlet['outlet_name'], f"{outlet['running_total']:,.2f}"))

    # Paginate 20 per page
    pdf = FPDF()
    pdf.set_font("Times", size=12)

    brand_name_out = brand_name
    if brand_name == 'Pukis & Martabak Kota Baru':
        brand_name_out = 'PKB'
    
    for page_start in range(1, len(table_rows), 20):
            pdf.add_page()
            if page_start == 1:
                # Only on the first page
                pdf.set_font("Times", style="B", size=16)
                pdf.cell(0, 10, "Outlet Net Income Rankings", ln=True, align="C")
                pdf.set_font("Times", style="", size=12)
                pdf.cell(0, 8, f"Brand: {brand_name_out}", ln=True, align="C")
                pdf.cell(0, 8, f"Period: {start_date.strftime('%d %b %Y')} - {end_date.strftime('%d %b %Y')}", ln=True, align="C")
                pdf.ln(2)
            else:
                # Add some spacing on subsequent pages to separate header from table
                pdf.ln(10)
            # pdf.set_font("Times", style="B", size=16)
            # pdf.cell(0, 10, "Outlet Net Income Rankings", ln=True, align="C")
            # pdf.set_font("Times", style="", size=12)
            # pdf.cell(0, 8, f"Brand: {brand_name}", ln=True, align="C")
            # pdf.cell(0, 8, f"Period: {start_date.strftime('%d %b %Y')} - {end_date.strftime('%d %b %Y')}", ln=True, align="C")
            # pdf.ln(2)
            with pdf.table(col_widths=(5,50,20), text_align=("center", "left", "left")) as table:
                # Always add header
                row = table.row()
                for datum in TABLE_HEADER:
                    row.cell(datum)
                # Add up to 20 rows
                for row_data in table_rows[page_start:page_start+20]:
                    row = table.row()
                    for datum in row_data:
                        row.cell(datum)
    pdf_bytes = BytesIO()
    pdf.output(pdf_bytes)
    pdf_bytes.seek(0)
    return send_file(pdf_bytes, mimetype='application/pdf', as_attachment=True, download_name= f'Top_Outlets_{brand_name_out}_{start_date.strftime("%Y%m%d")}_{end_date.strftime("%Y%m%d")}.pdf')


@reports_bp.route("/monthly-income", methods=["POST", "OPTIONS"])
@cross_origin(expose_headers=["Content-Disposition"])
def monthly_income_report():
    """
    Generates and returns an Excel report of monthly net income for a given brand.
    """
    # Request examples:
    # POST /reports/monthly-income?start_date=2025-12-01&end_date=2025-12-31 {"brand_name":"X"} -> 200
    # POST /reports/monthly-income?start_date=2025-12-01 {"brand_name":"X"} -> 400
    # POST /reports/monthly-income?start_date=2025-13-01&end_date=2025-12-31 {"brand_name":"X"} -> 400
    # POST /reports/monthly-income?start_date=2025-12-31&end_date=2025-12-01 {"brand_name":"X"} -> 400

    if request.method == 'OPTIONS':
        return jsonify({'status': 'OK'}), 200
    json_data = request.get_json()
    if not json_data:
        return jsonify({"error": "Invalid JSON"}), 400

    brand_name = json_data.get("brand_name")
    if not brand_name:
        return jsonify({"error": "brand_name is required"}), 400

    year = json_data.get("year", datetime.now().year)
    start_date, end_date, date_range_error = parse_date_range(request.args)
    if date_range_error:
        return jsonify({"error": date_range_error}), 400

    try:
        data = generate_monthly_net_income_data(
            brand_name,
            year,
            start_date=start_date,
            end_date=end_date,
        )
        if not data:
            return jsonify({"error": "No data found for the given criteria"}), 404

        workbook = openpyxl.Workbook()
        # Remove the default sheet created by openpyxl
        if "Sheet" in workbook.sheetnames:
            workbook.remove(workbook["Sheet"])

        sheet = MonthlyIncomeSheet(workbook, data)
        sheet.generate()

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        if start_date and end_date:
            download_name = (
                f"Monthly_income_{brand_name}_{start_date.isoformat()}_to_{end_date.isoformat()}.xlsx"
            )
        else:
            download_name = f"Monthly_income_{brand_name}_{year}.xlsx"

        response = send_file(
            output,
            as_attachment=True,
            download_name=download_name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        # ensure the browser can see Content-Disposition and set a safe referrer policy
        response.headers['Access-Control-Expose-Headers'] = 'Content-Disposition'
        response.headers['Referrer-Policy'] = 'no-referrer-when-downgrade'
        return response
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@reports_bp.route("/monthly-mpr-commission", methods=["POST", "OPTIONS"])
@cross_origin(expose_headers=["Content-Disposition"])
def monthly_mpr_commission_report():
    """
    Generates and returns an Excel report of monthly MPR commission totals.
    """
    if request.method == 'OPTIONS':
        return jsonify({'status': 'OK'}), 200

    json_data = request.get_json(silent=True) or {}
    year = json_data.get("year", datetime.now().year)
    start_date, end_date, date_range_error = parse_date_range(request.args)
    if date_range_error:
        return jsonify({"error": date_range_error}), 400

    try:
        data = generate_monthly_mpr_commission_data(
            year,
            start_date=start_date,
            end_date=end_date,
        )
        if not data:
            return jsonify({"error": "No data found for the given criteria"}), 404

        workbook = openpyxl.Workbook()
        if "Sheet" in workbook.sheetnames:
            workbook.remove(workbook["Sheet"])

        sheet = MonthlyMprCommissionSheet(workbook, data)
        sheet.generate()

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        if start_date and end_date:
            download_name = (
                f"Monthly_mpr_commission_MPR_{start_date.isoformat()}_to_{end_date.isoformat()}.xlsx"
            )
        else:
            download_name = f"Monthly_mpr_commission_MPR_{year}.xlsx"

        response = send_file(
            output,
            as_attachment=True,
            download_name=download_name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response.headers['Access-Control-Expose-Headers'] = 'Content-Disposition'
        response.headers['Referrer-Policy'] = 'no-referrer-when-downgrade'
        return response
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@reports_bp.route("/monthly-management-commission", methods=["POST", "OPTIONS"])
@cross_origin(expose_headers=["Content-Disposition"])
def monthly_management_commission_report():
    """
    Generates and returns an Excel report of monthly management commission totals.
    """
    if request.method == 'OPTIONS':
        return jsonify({'status': 'OK'}), 200

    json_data = request.get_json(silent=True) or {}
    brand_name = (json_data.get("brand_name") or "").strip()
    if not brand_name:
        return jsonify({"error": "brand_name is required"}), 400
    if brand_name.upper() == "MPR":
        return jsonify({"error": "brand_name must be a non-MPR brand"}), 400

    year = json_data.get("year", datetime.now().year)
    start_date, end_date, date_range_error = parse_date_range(request.args)
    if date_range_error:
        return jsonify({"error": date_range_error}), 400

    try:
        data = generate_monthly_management_commission_data(
            brand_name,
            year,
            start_date=start_date,
            end_date=end_date,
        )
        if not data or not data.get("outlets"):
            return jsonify({"error": "No data found for the given criteria"}), 404

        workbook = openpyxl.Workbook()
        if "Sheet" in workbook.sheetnames:
            workbook.remove(workbook["Sheet"])

        sheet = MonthlyManagementCommissionSheet(workbook, data)
        sheet.generate()

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        if start_date and end_date:
            download_name = (
                f"Monthly_management_commission_{brand_name}_{start_date.isoformat()}_to_{end_date.isoformat()}.xlsx"
            )
        else:
            download_name = f"Monthly_management_commission_{brand_name}_{year}.xlsx"

        response = send_file(
            output,
            as_attachment=True,
            download_name=download_name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response.headers['Access-Control-Expose-Headers'] = 'Content-Disposition'
        response.headers['Referrer-Policy'] = 'no-referrer-when-downgrade'
        return response
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@reports_bp.route("/monthly-management-commission-custom", methods=["POST", "OPTIONS"])
@cross_origin(expose_headers=["Content-Disposition"])
def monthly_management_commission_report_custom():
    """
    Generates an Excel report of monthly management commission totals using the
    literal start_date/end_date range from the request body.
    """
    if request.method == 'OPTIONS':
        return jsonify({'status': 'OK'}), 200

    json_data = request.get_json(silent=True) or {}
    brand_name = (json_data.get("brand_name") or "").strip()
    if not brand_name:
        return jsonify({"error": "brand_name is required"}), 400
    if brand_name.upper() == "MPR":
        return jsonify({"error": "brand_name must be a non-MPR brand"}), 400

    start_date, end_date, date_range_error = parse_date_range_from_body(json_data)
    if date_range_error:
        return jsonify({"error": date_range_error}), 400

    try:
        data = generate_monthly_management_commission_data_custom_range(
            brand_name,
            start_date,
            end_date,
        )
        if not data or not data.get("outlets"):
            return jsonify({"error": "No data found for the given criteria"}), 404

        workbook = openpyxl.Workbook()
        if "Sheet" in workbook.sheetnames:
            workbook.remove(workbook["Sheet"])

        sheet = MonthlyManagementCommissionSheet(workbook, data)
        sheet.generate()

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        download_name = (
            f"Monthly_management_commission_custom_{brand_name}_"
            f"{start_date.isoformat()}_to_{end_date.isoformat()}.xlsx"
        )

        response = send_file(
            output,
            as_attachment=True,
            download_name=download_name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response.headers['Access-Control-Expose-Headers'] = 'Content-Disposition'
        response.headers['Referrer-Policy'] = 'no-referrer-when-downgrade'
        return response
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# @reports_bp.route('/kas/report', methods=['GET'])
# def get_kas_report():
#     """
#     Generates and returns an Excel report of kas transactions.
#     """
#     start_date_param = request.args.get('start_date')
#     end_date_param = request.args.get('end_date')
#     report_type = request.args.get('type', 'Harian')

#     if not start_date_param or not end_date_param:
#         return jsonify({"error": "start_date and end_date are required"}), 400

#     try:
#         start_date = datetime.strptime(start_date_param, '%Y-%m-%d')
#         end_date = datetime.strptime(end_date_param, '%Y-%m-%d')

#         transactions = get_kas_transactions(start_date, end_date)

#         workbook = openpyxl.Workbook()
#         if "Sheet" in workbook.sheetnames:
#             workbook.remove(workbook["Sheet"])

#         data = {
#             'transactions': transactions,
#             'start_date': start_date,
#             'end_date': end_date,
#             'type': report_type
#         }

#         sheet = KasSheet(workbook, data)
#         sheet.generate()

#         output = io.BytesIO()
#         workbook.save(output)
#         output.seek(0)

#         return send_file(
#             output,
#             as_attachment=True,
#             download_name=f"Kas_{report_type}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx",
#             mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#         )
#     except Exception as e:
#         return jsonify({"error": str(e)}), 500

# @reports_bp.route('/kas', methods=['POST'])
# def add_kas_transaction():
#     """
#     Adds a new kas transaction.
#     """
#     data = request.get_json()
#     if not data or not all(k in data for k in ['tanggal', 'keterangan', 'tipe', 'jumlah']):
#         return jsonify({"error": "Missing data"}), 400
#     try:
#         new_transaction = create_kas_transaction(data)
#         return jsonify(new_transaction.to_dict()), 201
#     except Exception as e:
#         return jsonify({"error": str(e)}), 500

# @reports_bp.route('/kas/<int:transaction_id>', methods=['PATCH'])
# def edit_kas_transaction(transaction_id):
#     """
#     Updates a kas transaction.
#     """
#     data = request.get_json()
#     try:
#         updated_transaction = update_kas_transaction(transaction_id, data)
#         return jsonify(updated_transaction.to_dict()), 200
#     except Exception as e:
#         return jsonify({"error": str(e)}), 500

# @reports_bp.route('/kas/<int:transaction_id>', methods=['DELETE'])
# def remove_kas_transaction(transaction_id):
#     """
#     Deletes a kas transaction.
#     """
#     try:
#         delete_kas_transaction(transaction_id)
#         return jsonify({"message": "Transaction deleted"}), 200
#     except Exception as e:
#         return jsonify({"error": str(e)}), 500

# @reports_bp.route('/kas', methods=['GET'])
# def get_all_kas_transactions():
#     """
#     Retrieves all kas transactions.
#     """
#     try:
#         transactions = KasTransaction.query.all()
#         return jsonify([t.to_dict() for t in transactions]), 200
#     except Exception as e:
#         return jsonify({"error": str(e)}), 500

# @reports_bp.route('/kas/<int:transaction_id>', methods=['GET'])
# def get_kas_transaction_by_id(transaction_id):
#     """
#     Retrieves a single kas transaction by its ID.
#     """
#     try:
#         transaction = KasTransaction.query.get_or_404(transaction_id)
#         return jsonify(transaction.to_dict()), 200
#     except Exception as e:
#         return jsonify({"error": str(e)}), 500
