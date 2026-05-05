from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from app.models.mpr_mapping import MprMapping
from app.services.excel_export.base_sheet import BaseSheet
from app.services.excel_export import mpr_calculations as mpr_calc
from app.services.excel_export.utils.excel_utils import (
    HEADER_FONT, YELLOW_FILL, CENTER_ALIGN, GOJEK_FILL, GRAB_FILL, SHOPEE_FILL,
    SHOPEEPAY_FILL, TIKTOK_FILL, CASH_FILL, DATE_FILL, DIFFERENCE_FILL,
    set_column_widths
)

class DailySheet(BaseSheet):
    MPR_COMMISSION_RATE = 0.08
    OPTIONAL_MPR_AC_HEADERS = {
        'GoFood (ac)',
        'GO-PAY QRIS (ac)',
        'Gojek Net (ac)',
        'GrabFood (ac)',
        'GrabOVO (ac)',
        'Shopee Net (ac)',
        'ShopeePay Net (ac)',
        'Tiktok Net (ac)',
        'Qpon Net (ac)',
        'Webshop Net (ac)',
    }
    MP78_MANAGEMENT_AC_HEADERS = {
        'Gojek Net (ac)',
        'Grab Net (ac)',
        'Shopee Net (ac)',
        'ShopeePay Net (ac)',
        'Tiktok Net (ac)',
        'Qpon Net (ac)',
        'Webshop Net (ac)',
    }
    MP78_ENABLED_AC_HEADERS = {
        'Gojek Net (ac)',
        'Grab Net (ac)',
    }

    def __init__(self, workbook, data, sheet_name='Daily'):
        super().__init__(workbook, sheet_name, data)
        self._has_mpr_mapping = None

    def generate(self):
        self._write_title()
        self._write_headers()
        self._write_data()
        self._write_grand_total()
        self._set_column_widths()

    def _write_title(self):
        self.ws['A1'] = 'Sales Report'
        self.ws['A2'] = 'Period:'
        self.ws['B2'] = f"{self.data['start_date'].strftime('%Y-%m-%d')} to {self.data['end_date'].strftime('%Y-%m-%d')}"
        self.ws['A3'] = 'Outlet:'
        self.ws['B3'] = self.data['outlet'].outlet_name_gojek

    def _get_headers(self):
        user_role = self.data.get('user_role')
        outlet_brand = self.data['outlet'].brand

        base_headers = [
            'Date', 'GoFood', 'GO-PAY QRIS', 'Gojek Net', 'Gojek Mutation', 'Gojek Difference', 'Gojek Net (ac)',
            'GrabFood', 'GrabOVO', 'Grab Net (ac)', 'Shopee Net', 'Shopee Mutation', 'Shopee Difference',
            'ShopeePay Net', 'ShopeePay Mutation', 'ShopeePay Difference', 'ShopeePay Net (ac)',
            'Tiktok Net', 'Qpon Net', 'Webshop Net', 'UV'
        ]
        extra_headers = ['Cash Income (Admin)', 'Cash Expense (Admin)', 'Sisa Cash (Admin)', 'Minusan (Mutasi)']

        if outlet_brand == 'MPR':
            base_headers = [
                'Date',
                'GoFood', 'GO-PAY QRIS', 'GoFood (ac)', 'GO-PAY QRIS (ac)',
                'Gojek Net', 'Gojek Mutation', 'Gojek Difference', 'Gojek Net (ac)',
                'GrabFood', 'GrabOVO', 'GrabFood (ac)', 'GrabOVO (ac)', 'Grab Net', 'Grab Net (ac)',
                'Shopee Net', 'Shopee Mutation', 'Shopee Difference', 'Shopee Net (ac)',
                'ShopeePay Net', 'ShopeePay Mutation', 'ShopeePay Difference', 'ShopeePay Net (ac)',
                'Tiktok Net', 'Tiktok Net (ac)',
                'Qpon Net', 'Qpon Net (ac)',
                'Webshop Net', 'Webshop Net (ac)',
                'UV'
            ]

        if (
            user_role == "management"
            and outlet_brand not in ["MPR", "Pukis & Martabak Kota Baru", "Es Ce Hun Tiau & Bongko Wendy"]
        ):
            base_headers.insert(8, 'Grab Net')

        if outlet_brand == "Pukis & Martabak Kota Baru" or outlet_brand == "Es Ce Hun Tiau & Bongko Wendy":
            base_headers.insert(4, 'Grab Net')
            if 'Grab Net (ac)' in base_headers:
                base_headers.remove('Grab Net (ac)')
            base_headers += extra_headers
        if outlet_brand == "Es Ce Hun Tiau & Bongko Wendy":
            base_headers.insert(9, 'Grab Net (ac)')
            if 'Grab Net' in base_headers:
                base_headers.remove('Grab Net')

        if outlet_brand == 'MP78':
            if mpr_calc.ENABLE_MP78_MANAGEMENT_AC:
                base_headers = self._add_mp78_management_ac_headers(base_headers)
                base_headers = [
                    header for header in base_headers
                    if (
                        header not in self.MP78_MANAGEMENT_AC_HEADERS
                        or header in self.MP78_ENABLED_AC_HEADERS
                    )
                ]
            else:
                base_headers = [
                    header for header in base_headers
                    if header not in self.MP78_MANAGEMENT_AC_HEADERS
                ]

        if outlet_brand == 'MPR' and not self._current_outlet_has_mpr_mapping():
            base_headers = [
                header for header in base_headers
                if header not in self.OPTIONAL_MPR_AC_HEADERS
            ]

        return base_headers

    def _add_mp78_management_ac_headers(self, headers):
        header_positions = [
            ('Shopee Difference', 'Shopee Net (ac)'),
            ('Tiktok Net', 'Tiktok Net (ac)'),
            ('Qpon Net', 'Qpon Net (ac)'),
            ('Webshop Net', 'Webshop Net (ac)'),
        ]

        for anchor, header in header_positions:
            if header not in headers and anchor in headers:
                headers.insert(headers.index(anchor) + 1, header)

        return headers

    def _get_value_with_mutation_fallback(self, totals, mutation_key, net_key):
        return mpr_calc.value_with_mutation_fallback(totals, mutation_key, net_key)

    def _get_mpr_adjusted_value(self, totals, net_key, mutation_key=None):
        display_value = (
            self._get_value_with_mutation_fallback(totals, mutation_key, net_key)
            if mutation_key else totals.get(net_key, 0)
        )
        return display_value - (self.MPR_COMMISSION_RATE * totals.get(net_key, 0))

    def _is_mpr_brand(self):
        return self.data['outlet'].brand == 'MPR'

    def _is_mp78_brand(self):
        return self.data['outlet'].brand == 'MP78'

    def _uses_mp78_management_ac(self):
        return self._is_mp78_brand() and mpr_calc.ENABLE_MP78_MANAGEMENT_AC

    def _get_gofood_value(self, totals):
        return mpr_calc.gofood_value(totals)

    def _get_gofood_ac_value(self, totals):
        return mpr_calc.gofood_value(totals, self._is_mpr_brand())

    def _get_gojek_qris_value(self, totals):
        return mpr_calc.gojek_qris_value(totals)

    def _get_gojek_qris_ac_value(self, totals):
        return mpr_calc.gojek_qris_value(totals, self._is_mpr_brand())

    def _get_gojek_net_value(self, totals):
        return mpr_calc.gojek_net_value(totals, self._is_mpr_brand())

    def _get_gojek_net_ac_value(self, totals):
        if self._is_mpr_brand() and self._current_outlet_has_mpr_mapping():
            return mpr_calc.gojek_net_ac_value(totals)

        if self._uses_mp78_management_ac():
            return mpr_calc.mp78_ac_value_for_header(totals, 'Gojek_Mutation')

        return self._get_value_with_mutation_fallback(totals, 'Gojek_Mutation', 'Gojek_Net')

    def _get_grabfood_value(self, totals):
        return mpr_calc.grabfood_value(totals)

    def _get_grabfood_ac_value(self, totals):
        return mpr_calc.grabfood_value(totals, self._is_mpr_brand())

    def _get_grab_ovo_value(self, totals):
        return mpr_calc.grab_ovo_value(totals)

    def _get_grab_ovo_ac_value(self, totals):
        return mpr_calc.grab_ovo_value(totals, self._is_mpr_brand())

    def _get_grab_net_value(self, totals):
        return mpr_calc.grab_net_value(totals, self._is_mpr_brand())

    def _get_shopee_net_value(self, totals):
        return mpr_calc.shopee_net_value(totals)

    def _get_shopee_net_ac_value(self, totals):
        if self._is_mpr_brand() and self._current_outlet_has_mpr_mapping():
            return mpr_calc.shopee_net_ac_value(totals)

        if self._uses_mp78_management_ac():
            return mpr_calc.mp78_ac_value_for_header(totals, 'Shopee_Net')

        return self._get_value_with_mutation_fallback(totals, 'Shopee_Mutation', 'Shopee_Net')

    def _get_shopeepay_net_value(self, totals):
        return mpr_calc.shopeepay_net_value(totals)

    def _get_shopeepay_net_ac_value(self, totals):
        if self._is_mpr_brand() and self._current_outlet_has_mpr_mapping():
            return mpr_calc.shopeepay_net_ac_value(totals)

        if self._uses_mp78_management_ac():
            return mpr_calc.mp78_ac_value_for_header(totals, 'ShopeePay_Net')

        return self._get_value_with_mutation_fallback(totals, 'ShopeePay_Mutation', 'ShopeePay_Net')

    def _current_outlet_has_mpr_mapping(self):
        if self._has_mpr_mapping is None:
            outlet_code = self.data['outlet'].outlet_code
            self._has_mpr_mapping = (
                MprMapping.query.filter_by(mpr_outlet_code=outlet_code).first() is not None
            )

        return self._has_mpr_mapping

    def _get_grab_net_ac_value(self, totals):
        if (
            self._is_mpr_brand()
            and self._current_outlet_has_mpr_mapping()
        ):
            return mpr_calc.grab_net_ac_value(totals)

        if self._uses_mp78_management_ac():
            return mpr_calc.mp78_ac_value_for_header(totals, 'Grab_Net')

        return mpr_calc.management_net_ac_value(totals, 'Grab_Net')

    def _get_standard_net_ac_value(self, totals, net_key):
        if self._uses_mp78_management_ac():
            return mpr_calc.mp78_ac_value_for_header(totals, net_key)

        return self._get_mpr_adjusted_value(totals, net_key)

    def _set_column_widths(self):
        widths = {
            get_column_letter(col): 15
            for col in range(1, self.ws.max_column + 1)
        }
        widths['A'] = 12
        set_column_widths(self.ws, widths)

    def _write_headers(self):
        headers = self._get_headers()
        header_colors = {
            'Date': DATE_FILL, 'GoFood': GOJEK_FILL, 'GO-PAY QRIS': GOJEK_FILL, 'GoFood (ac)': GOJEK_FILL,
            'GO-PAY QRIS (ac)': GOJEK_FILL, 'Gojek Net': GOJEK_FILL, 'Gojek Mutation': GOJEK_FILL,
            'Gojek Difference': DIFFERENCE_FILL, 'GrabFood': GRAB_FILL, 'GrabOVO': GRAB_FILL,
            'GrabFood (ac)': GRAB_FILL, 'GrabOVO (ac)': GRAB_FILL, 'Grab Net': GRAB_FILL, 'Grab Net (ac)': GRAB_FILL,
            'Gojek Net (ac)': GOJEK_FILL, 'Grab Net (ac)': GRAB_FILL,
            'Shopee Net (ac)': SHOPEE_FILL, 'ShopeePay Net (ac)': SHOPEEPAY_FILL,
            'Tiktok Net (ac)': TIKTOK_FILL, 'Qpon Net (ac)': TIKTOK_FILL,
            'Webshop Net (ac)': TIKTOK_FILL,
            'Shopee Net': SHOPEE_FILL, 'Shopee Mutation': SHOPEE_FILL, 'Shopee Difference': DIFFERENCE_FILL,
            'ShopeePay Net': SHOPEEPAY_FILL, 'ShopeePay Mutation': SHOPEEPAY_FILL, 'ShopeePay Difference': DIFFERENCE_FILL,
            'Tiktok Net': TIKTOK_FILL, 'Qpon Net': TIKTOK_FILL, 'Webshop Net': TIKTOK_FILL,
            'UV': PatternFill(start_color='35F0F0', end_color='35F0F0', fill_type='solid'),
            'Cash Income (Admin)': CASH_FILL, 'Cash Expense (Admin)': CASH_FILL, 'Sisa Cash (Admin)': CASH_FILL
        }

        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=5, column=col)
            cell.value = header
            cell.font = HEADER_FONT
            cell.fill = header_colors.get(header, YELLOW_FILL)
            cell.alignment = CENTER_ALIGN

    def _write_data(self):
        headers = self._get_headers()
        all_dates = self.data['all_dates']
        daily_totals = self.data['daily_totals']
        minusan_by_date = self.data['minusan_by_date']
        current_row = 6

        header_value_map = {
            'Date': lambda totals, date, minusan_total: date,
            'GoFood': lambda totals, date, minusan_total: self._get_gofood_value(totals),
            'GO-PAY QRIS': lambda totals, date, minusan_total: self._get_gojek_qris_value(totals),
            'GoFood (ac)': lambda totals, date, minusan_total: self._get_gofood_ac_value(totals),
            'GO-PAY QRIS (ac)': lambda totals, date, minusan_total: self._get_gojek_qris_ac_value(totals),
            'Gojek Net': lambda totals, date, minusan_total: self._get_gojek_net_value(totals),
            'Gojek Mutation': lambda totals, date, minusan_total: totals.get('Gojek_Mutation', 0),
            'Gojek Net (ac)': lambda totals, date, minusan_total: self._get_gojek_net_ac_value(totals),
            'Gojek Difference': lambda totals, date, minusan_total: totals.get('Gojek_Difference', 0),
            'GrabFood': lambda totals, date, minusan_total: self._get_grabfood_value(totals),
            'GrabOVO': lambda totals, date, minusan_total: self._get_grab_ovo_value(totals),
            'GrabFood (ac)': lambda totals, date, minusan_total: self._get_grabfood_ac_value(totals),
            'GrabOVO (ac)': lambda totals, date, minusan_total: self._get_grab_ovo_ac_value(totals),
            'Grab Net': lambda totals, date, minusan_total: self._get_grab_net_value(totals),
            'Grab Net (ac)': lambda totals, date, minusan_total: self._get_grab_net_ac_value(totals),
            'Shopee Net': lambda totals, date, minusan_total: self._get_shopee_net_value(totals),
            'Shopee Mutation': lambda totals, date, minusan_total: totals.get('Shopee_Mutation', 0),
            'Shopee Net (ac)': lambda totals, date, minusan_total: self._get_shopee_net_ac_value(totals),
            'Shopee Difference': lambda totals, date, minusan_total: totals.get('Shopee_Difference', 0),
            'ShopeePay Net': lambda totals, date, minusan_total: self._get_shopeepay_net_value(totals),
            'ShopeePay Mutation': lambda totals, date, minusan_total: totals.get('ShopeePay_Mutation', 0),
            'ShopeePay Net (ac)': lambda totals, date, minusan_total: self._get_shopeepay_net_ac_value(totals),
            'ShopeePay Difference': lambda totals, date, minusan_total: totals.get('ShopeePay_Difference', 0),
            'Tiktok Net': lambda totals, date, minusan_total: totals.get('Tiktok_Net', 0),
            'Tiktok Net (ac)': lambda totals, date, minusan_total: self._get_standard_net_ac_value(totals, 'Tiktok_Net'),
            'Qpon Net': lambda totals, date, minusan_total: totals.get('Qpon_Net', 0),
            'Qpon Net (ac)': lambda totals, date, minusan_total: self._get_standard_net_ac_value(totals, 'Qpon_Net'),
            'Webshop Net': lambda totals, date, minusan_total: totals.get('Webshop_Net', 0),
            'Webshop Net (ac)': lambda totals, date, minusan_total: self._get_standard_net_ac_value(totals, 'Webshop_Net'),
            'UV': lambda totals, date, minusan_total: totals.get('UV', 0),
            'Cash Income (Admin)': lambda totals, date, minusan_total: totals.get('Cash_Income', 0),
            'Cash Expense (Admin)': lambda totals, date, minusan_total: totals.get('Cash_Expense', 0),
            'Sisa Cash (Admin)': lambda totals, date, minusan_total: totals.get('Cash_Income', 0) - totals.get('Cash_Expense', 0),
            'Minusan (Mutasi)': lambda totals, date, minusan_total: minusan_total,
        }

        for date in all_dates:
            totals = daily_totals.get(date, {})
            minusan_total = minusan_by_date.get(date, 0)
            row_data = [
                header_value_map[header](totals, date, minusan_total) if header in header_value_map else None
                for header in headers
            ]

            for col, value in enumerate(row_data, 1):
                cell = self.ws.cell(row=current_row, column=col, value=value)
                if isinstance(value, (int, float)) and col > 1:
                    cell.number_format = '#,##0'

                # Apply conditional formatting for difference columns
                if headers[col-1] in ['Gojek Difference', 'Shopee Difference', 'ShopeePay Difference']:
                    if value is not None:
                        if value > 0:
                            cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Light green
                        elif value < 0:
                            cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # Light red
            current_row += 1

    def _write_grand_total(self):
        headers = self._get_headers()
        grand_totals = self.data['grand_totals']
        all_dates = self.data['all_dates']
        minusan_by_date = self.data['minusan_by_date']
        grand_total_row = self.ws.max_row + 1

        grand_total_value_map = {
            'Date': lambda: 'Grand Total',
            'GoFood': lambda: self._get_gofood_value(grand_totals),
            'GO-PAY QRIS': lambda: self._get_gojek_qris_value(grand_totals),
            'GoFood (ac)': lambda: self._get_gofood_ac_value(grand_totals),
            'GO-PAY QRIS (ac)': lambda: self._get_gojek_qris_ac_value(grand_totals),
            'Gojek Net': lambda: self._get_gojek_net_value(grand_totals),
            'Gojek Mutation': lambda: grand_totals.get('Gojek_Mutation', 0),
            'Gojek Net (ac)': lambda: self._get_gojek_net_ac_value(grand_totals),
            'Gojek Difference': lambda: grand_totals.get('Gojek_Difference', 0),
            'GrabFood': lambda: self._get_grabfood_value(grand_totals),
            'GrabOVO': lambda: self._get_grab_ovo_value(grand_totals),
            'GrabFood (ac)': lambda: self._get_grabfood_ac_value(grand_totals),
            'GrabOVO (ac)': lambda: self._get_grab_ovo_ac_value(grand_totals),
            'Grab Net': lambda: self._get_grab_net_value(grand_totals),
            'Grab Net (ac)': lambda: self._get_grab_net_ac_value(grand_totals),
            'Shopee Net': lambda: self._get_shopee_net_value(grand_totals),
            'Shopee Mutation': lambda: grand_totals.get('Shopee_Mutation', 0),
            'Shopee Net (ac)': lambda: self._get_shopee_net_ac_value(grand_totals),
            'Shopee Difference': lambda: grand_totals.get('Shopee_Difference', 0),
            'ShopeePay Net': lambda: self._get_shopeepay_net_value(grand_totals),
            'ShopeePay Mutation': lambda: grand_totals.get('ShopeePay_Mutation', 0),
            'ShopeePay Net (ac)': lambda: self._get_shopeepay_net_ac_value(grand_totals),
            'ShopeePay Difference': lambda: grand_totals.get('ShopeePay_Difference', 0),
            'Tiktok Net': lambda: grand_totals.get('Tiktok_Net', 0),
            'Tiktok Net (ac)': lambda: self._get_standard_net_ac_value(grand_totals, 'Tiktok_Net'),
            'Qpon Net': lambda: grand_totals.get('Qpon_Net', 0),
            'Qpon Net (ac)': lambda: self._get_standard_net_ac_value(grand_totals, 'Qpon_Net'),
            'Webshop Net': lambda: grand_totals.get('Webshop_Net', 0),
            'Webshop Net (ac)': lambda: self._get_standard_net_ac_value(grand_totals, 'Webshop_Net'),
            'UV': lambda: grand_totals.get('UV', 0),
            'Cash Income (Admin)': lambda: grand_totals.get('Cash_Income', 0),
            'Cash Expense (Admin)': lambda: grand_totals.get('Cash_Expense', 0),
            'Sisa Cash (Admin)': lambda: grand_totals.get('Cash_Difference', 0),
            'Minusan (Mutasi)': lambda: sum(minusan_by_date.get(date, 0) for date in all_dates),
        }

        row_data = [
            grand_total_value_map[header]() if header in grand_total_value_map else None
            for header in headers
        ]

        for col, value in enumerate(row_data, 1):
            cell = self.ws.cell(row=grand_total_row, column=col, value=value)
            cell.font = HEADER_FONT
            cell.fill = YELLOW_FILL
            cell.alignment = CENTER_ALIGN
            if isinstance(value, (int, float)):
                cell.number_format = '#,##0'
