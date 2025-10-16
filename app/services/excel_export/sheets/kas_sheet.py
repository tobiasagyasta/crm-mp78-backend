from openpyxl.styles import Font, Alignment
from app.services.excel_export.base_sheet import BaseSheet

class KasSheet(BaseSheet):
    def __init__(self, workbook, data, sheet_name='Kas'):
        super().__init__(workbook, sheet_name, data)

    def generate(self):
        self._write_title()
        self._write_headers()
        self._write_data()

    def _write_title(self):
        start_date = self.data['start_date'].strftime('%d/%m/%Y')
        end_date = self.data['end_date'].strftime('%d/%m/%Y')
        title = f"Kas {self.data['type']} ({start_date} – {end_date})"
        self.ws.merge_cells('A1:E1')
        cell = self.ws['A1']
        cell.value = title
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal='center')

    def _write_headers(self):
        headers = ['Tanggal', 'Keterangan', 'Uang Masuk', 'Uang Keluar', 'Sisa Saldo']
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

    def _write_data(self):
        transactions = self.data['transactions']
        balance = 0
        current_row = 4

        for trx in transactions:
            balance += trx.jumlah if trx.tipe == 'Masuk' else -trx.jumlah

            self.ws.cell(row=current_row, column=1, value=trx.tanggal)
            self.ws.cell(row=current_row, column=2, value=trx.keterangan)

            if trx.tipe == 'Masuk':
                self.ws.cell(row=current_row, column=3, value=trx.jumlah)
            else:
                self.ws.cell(row=current_row, column=4, value=trx.jumlah)

            self.ws.cell(row=current_row, column=5, value=balance)

            # Apply currency format
            self.ws.cell(row=current_row, column=3).number_format = '#,##0.00'
            self.ws.cell(row=current_row, column=4).number_format = '#,##0.00'
            self.ws.cell(row=current_row, column=5).number_format = '#,##0.00'

            # Apply date format
            self.ws.cell(row=current_row, column=1).number_format = 'DD/MM/YY'

            current_row += 1