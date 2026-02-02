import calendar
from openpyxl.styles import Alignment, Font, NamedStyle
from openpyxl.utils import get_column_letter

from app.services.excel_export.base_sheet import BaseSheet


class MonthlyIncomeSheet(BaseSheet):
    def __init__(self, workbook, data):
        super().__init__(workbook, "Monthly Net Income", data)

    def generate(self):
        """Generates the monthly net income Excel sheet."""
        self._prepare_styles()
        self._create_header()
        self._populate_data()
        self._set_column_widths()

    def _prepare_styles(self):
        self.header_font = Font(bold=True)
        self.center_align = Alignment(horizontal="center", vertical="center")
        self.number_style = NamedStyle(name='number_style', number_format='#,##0')
        if 'number_style' not in self.wb.style_names:
            self.wb.add_named_style(self.number_style)

    def _create_header(self):
        if isinstance(self.data, dict) and 'periods' in self.data:
            periods = self.data['periods']
            period_headers = [f"{calendar.month_name[month]} {year}" for year, month in periods]
            headers = ['Outlet', 'Closing Day'] + period_headers + ['Total Per Outlet']
        else:
            headers = ['Outlet', 'Closing Day'] + [calendar.month_name[i] for i in range(1, 13)] + ['Total Per Outlet']
        self.ws.append(headers)
        for cell in self.ws[1]:
            cell.font = self.header_font
            cell.alignment = self.center_align

    def _populate_data(self):
        if isinstance(self.data, dict) and 'periods' in self.data:
            periods = self.data['periods']
            outlets = self.data.get('outlets', {})
            column_totals = {period: 0 for period in periods}
            grand_total = 0
            for outlet_code, outlet_data in outlets.items():
                monthly_total = outlet_data.get(
                    'total',
                    sum(outlet_data['monthly_totals'].get(period, 0) for period in periods),
                )
                row_data = [
                    outlet_data['name'],
                    outlet_data['closing_day'],
                ]
                for period in periods:
                    period_value = outlet_data['monthly_totals'].get(period, 0)
                    column_totals[period] += period_value
                    row_data.append(period_value)
                row_data.append(monthly_total)

                self.ws.append(row_data)

                # Apply styles
                self.ws.cell(row=self.ws.max_row, column=2).alignment = self.center_align
                for col_num in range(3, 3 + len(periods) + 1):
                    self.ws.cell(row=self.ws.max_row, column=col_num).style = 'number_style'
                grand_total += monthly_total

            total_row = ['Total Per Month', '']
            for period in periods:
                total_row.append(column_totals.get(period, 0))
            total_row.append(grand_total)
            self.ws.append(total_row)
            for col_num in range(1, 3 + len(periods) + 1):
                cell = self.ws.cell(row=self.ws.max_row, column=col_num)
                cell.font = self.header_font
                if col_num >= 3:
                    cell.style = 'number_style'
        else:
            column_totals = {i: 0 for i in range(1, 13)}
            grand_total = 0
            for outlet_code, outlet_data in self.data.items():
                monthly_total = outlet_data.get(
                    'total',
                    sum(outlet_data['monthly_totals'].get(i, 0) for i in range(1, 13)),
                )
                row_data = [
                    outlet_data['name'],
                    outlet_data['closing_day'],
                ]
                for i in range(1, 13):
                    month_value = outlet_data['monthly_totals'].get(i, 0)
                    column_totals[i] += month_value
                    row_data.append(month_value)
                row_data.append(monthly_total)

                self.ws.append(row_data)

                # Apply styles
                self.ws.cell(row=self.ws.max_row, column=2).alignment = self.center_align
                for col_num in range(3, 16):
                    self.ws.cell(row=self.ws.max_row, column=col_num).style = 'number_style'
                grand_total += monthly_total

            total_row = ['Total Per Month', '']
            for i in range(1, 13):
                total_row.append(column_totals.get(i, 0))
            total_row.append(grand_total)
            self.ws.append(total_row)
            for col_num in range(1, 16):
                cell = self.ws.cell(row=self.ws.max_row, column=col_num)
                cell.font = self.header_font
                if col_num >= 3:
                    cell.style = 'number_style'

    def _set_column_widths(self):
        self.ws.column_dimensions[get_column_letter(1)].width = 35
        self.ws.column_dimensions[get_column_letter(2)].width = 15
        if isinstance(self.data, dict) and 'periods' in self.data:
            column_count = 3 + len(self.data['periods'])
        else:
            column_count = 15
        for i in range(3, column_count + 1):
            self.ws.column_dimensions[get_column_letter(i)].width = 15
