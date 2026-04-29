import calendar
from datetime import date, datetime, time, timedelta

from flask import has_app_context, has_request_context, request
from openpyxl.styles import Alignment, Font, NamedStyle, PatternFill
from openpyxl.utils import get_column_letter

from app.models.qpon_reports import QponReport
from app.models.tiktok_reports import TiktokReport
from app.models.webshop_report import WebshopReport
from app.services.excel_export.base_sheet import BaseSheet


class MonthlyManagementCommissionSheet(BaseSheet):
    PLATFORM_GROUPS = (
        {
            "key": "grab",
            "label": "Grab",
            "net_key": "net_total",
            "commission_key": "commission_total",
            "after_key": "net_after_commission",
            "rate": 1 / 74,
        },
        {
            "key": "tiktok",
            "label": "TikTok",
            "net_key": "tiktok_net",
            "commission_key": "tiktok_commission",
            "after_key": "tiktok_net_after_commission",
            "rate": 1 / 74,
        },
        {
            "key": "qpon",
            "label": "Qpon",
            "net_key": "qpon_net",
            "commission_key": "qpon_commission",
            "after_key": "qpon_net_after_commission",
            "rate": 1 / 74,
        },
        {
            "key": "webshop",
            "label": "Webshop",
            "net_key": "webshop_net",
            "commission_key": "webshop_commission",
            "after_key": "webshop_net_after_commission",
            "rate": 0.03,
        },
    )
    PERIOD_COLUMN_SPAN = len(PLATFORM_GROUPS) * 3

    def __init__(self, workbook, data):
        super().__init__(workbook, "Monthly Management Commission", data)

    def generate(self):
        self._prepare_styles()
        self._create_header()
        self._populate_data()
        self._set_column_widths()
        self.ws.freeze_panes = "C3"

    def _prepare_styles(self):
        self.header_font = Font(bold=True)
        self.center_align = Alignment(horizontal="center", vertical="center")
        self.header_fill = PatternFill(
            start_color="D9EAD3",
            end_color="D9EAD3",
            fill_type="solid",
        )
        self.commission_fill = PatternFill(
            start_color="FFF2CC",
            end_color="FFF2CC",
            fill_type="solid",
        )
        self.net_after_fill = PatternFill(
            start_color="DDEBF7",
            end_color="DDEBF7",
            fill_type="solid",
        )
        self.currency_style = NamedStyle(
            name="management_commission_currency_style",
            number_format="#,##0.00",
        )
        if "management_commission_currency_style" not in self.wb.style_names:
            self.wb.add_named_style(self.currency_style)

    def _periods(self):
        return self.data.get("periods", list(range(1, 13)))

    def _outlets(self):
        return self.data.get("outlets", {})

    def _report_year(self):
        if self.data.get("year"):
            return int(self.data["year"])

        if has_request_context():
            request_data = request.get_json(silent=True) or {}
            if request_data.get("year"):
                return int(request_data["year"])

        return None

    def _period_label(self, period):
        if isinstance(period, tuple):
            year, month = period
            return f"{calendar.month_name[month]} {year}"
        return calendar.month_name[period]

    def _parse_opening_day(self, closing_date_str):
        if not closing_date_str or closing_date_str == "Calendar" or "-" not in closing_date_str:
            return None
        try:
            start_day = int(str(closing_date_str).split("-")[0])
            if 1 <= start_day <= 31:
                return start_day
        except (ValueError, TypeError, IndexError):
            return None
        return None

    def _next_month(self, period_year, period_month):
        if period_month == 12:
            return period_year + 1, 1
        return period_year, period_month + 1

    def _build_period_window(self, period, closing_day):
        if isinstance(period, tuple):
            period_year, period_month = period
        else:
            report_year = self._report_year()
            if not report_year:
                return None
            period_year, period_month = report_year, period

        opening_day = self._parse_opening_day(closing_day)
        if not opening_day:
            period_start = date(period_year, period_month, 1)
            next_year, next_month = self._next_month(period_year, period_month)
            period_end = date(next_year, next_month, 1) - timedelta(days=1)
            return period_start, period_end

        next_year, next_month = self._next_month(period_year, period_month)
        period_start = date(period_year, period_month, 1) + timedelta(days=opening_day - 1)
        period_end = date(next_year, next_month, 1) + timedelta(days=opening_day - 2)
        return period_start, period_end

    def _resolve_period(self, transaction_date, periods, closing_day):
        opening_day = self._parse_opening_day(closing_day)
        financial_year = transaction_date.year
        financial_month = transaction_date.month

        if opening_day and transaction_date.day < opening_day:
            financial_month -= 1
            if financial_month == 0:
                financial_month = 12
                financial_year -= 1

        if periods and isinstance(periods[0], tuple):
            return financial_year, financial_month
        return financial_month

    def _query_platform_period_nets(self, outlet_code, periods, closing_day):
        empty_platform_nets = {period: {} for period in periods}
        if not has_app_context():
            return empty_platform_nets

        period_windows = {
            period: self._build_period_window(period, closing_day)
            for period in periods
        }
        valid_windows = [window for window in period_windows.values() if window]
        if not outlet_code or not valid_windows:
            return empty_platform_nets

        query_start = min(window[0] for window in valid_windows)
        query_end = max(window[1] for window in valid_windows)
        start_dt = datetime.combine(query_start, time.min)
        end_dt = datetime.combine(query_end, time.max)

        platform_nets = {
            period: {"tiktok": 0, "qpon": 0, "webshop": 0}
            for period in periods
        }
        period_set = set(periods)

        for report in TiktokReport.query.filter(
            TiktokReport.outlet_code == outlet_code,
            TiktokReport.order_time >= start_dt,
            TiktokReport.order_time <= end_dt,
        ).all():
            if not report.order_time:
                continue
            period = self._resolve_period(report.order_time.date(), periods, closing_day)
            if period in period_set:
                platform_nets[period]["tiktok"] += float(report.net_amount or 0)

        for report in QponReport.query.filter(
            QponReport.outlet_code == outlet_code,
            QponReport.bill_created_at >= start_dt,
            QponReport.bill_created_at <= end_dt,
        ).all():
            if not report.bill_created_at:
                continue
            period = self._resolve_period(report.bill_created_at.date(), periods, closing_day)
            if period in period_set:
                platform_nets[period]["qpon"] += float(report.nett_amount or 0)

        for report in WebshopReport.query.filter(
            WebshopReport.outlet_code == outlet_code,
            WebshopReport.created_at >= start_dt,
            WebshopReport.created_at <= end_dt,
        ).all():
            if not report.created_at:
                continue
            period = self._resolve_period(report.created_at.date(), periods, closing_day)
            if period in period_set:
                platform_nets[period]["webshop"] += float(report.nett_value or 0)

        return platform_nets

    def _platform_values(self, period_totals, platform, fallback_nets):
        net_total = float(
            period_totals.get(
                platform["net_key"],
                fallback_nets.get(platform["key"], 0),
            )
            or 0
        )
        commission_total = float(
            period_totals.get(
                platform["commission_key"],
                net_total * platform["rate"],
            )
            or 0
        )
        net_after_commission = float(
            period_totals.get(
                platform["after_key"],
                net_total - commission_total,
            )
            or 0
        )
        return net_total, commission_total, net_after_commission

    def _create_header(self):
        periods = self._periods()

        self.ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        self.ws["A1"] = "Outlet"

        self.ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
        self.ws["B1"] = "Closing Date"

        current_col = 3
        for period in periods:
            self.ws.merge_cells(
                start_row=1,
                start_column=current_col,
                end_row=1,
                end_column=current_col + self.PERIOD_COLUMN_SPAN - 1,
            )
            self.ws.cell(row=1, column=current_col, value=self._period_label(period))
            for platform_index, platform in enumerate(self.PLATFORM_GROUPS):
                platform_col = current_col + (platform_index * 3)
                self.ws.cell(row=2, column=platform_col, value=f"{platform['label']} Net")
                self.ws.cell(row=2, column=platform_col + 1, value="Commission")
                self.ws.cell(
                    row=2,
                    column=platform_col + 2,
                    value=f"{platform['label']} Net After Commission",
                )
            current_col += self.PERIOD_COLUMN_SPAN

        self.ws.merge_cells(
            start_row=1,
            start_column=current_col,
            end_row=2,
            end_column=current_col,
        )
        self.ws.cell(row=1, column=current_col, value="Total After Commission")

        for row in self.ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=current_col):
            for cell in row:
                cell.font = self.header_font
                cell.alignment = self.center_align
                cell.fill = self.header_fill

        for col_num in range(3, current_col):
            position = (col_num - 3) % 3
            if position == 1:
                self.ws.cell(row=2, column=col_num).fill = self.commission_fill
            elif position == 2:
                self.ws.cell(row=2, column=col_num).fill = self.net_after_fill
        self.ws.cell(row=1, column=current_col).fill = self.net_after_fill

    def _populate_data(self):
        periods = self._periods()
        outlets = self._outlets()
        period_totals_by_platform = {
            period: {
                platform["key"]: {"net": 0, "commission": 0, "after": 0}
                for platform in self.PLATFORM_GROUPS
            }
            for period in periods
        }
        grand_total_after_commission = 0
        current_row = 3

        for outlet_code, outlet_data in outlets.items():
            row_values = [
                outlet_data.get("name", ""),
                outlet_data.get("closing_day", "Calendar"),
            ]
            platform_period_nets = self._query_platform_period_nets(
                outlet_code,
                periods,
                outlet_data.get("closing_day", "Calendar"),
            )
            total_after_commission = 0

            for period in periods:
                period_totals = outlet_data.get("monthly_totals", {}).get(
                    period,
                    {},
                )
                if not isinstance(period_totals, dict):
                    period_totals = {"net_total": period_totals}
                fallback_nets = platform_period_nets.get(period, {})
                for platform in self.PLATFORM_GROUPS:
                    net_total, commission_total, net_after_commission = self._platform_values(
                        period_totals,
                        platform,
                        fallback_nets,
                    )
                    period_totals_by_platform[period][platform["key"]]["net"] += net_total
                    period_totals_by_platform[period][platform["key"]]["commission"] += (
                        commission_total
                    )
                    period_totals_by_platform[period][platform["key"]]["after"] += (
                        net_after_commission
                    )
                    total_after_commission += net_after_commission
                    row_values.extend([net_total, commission_total, net_after_commission])

            row_values.append(total_after_commission)
            self.ws.append(row_values)

            self.ws.cell(row=current_row, column=2).alignment = self.center_align
            for col_num in range(3, len(row_values) + 1):
                self.ws.cell(row=current_row, column=col_num).style = (
                    "management_commission_currency_style"
                )
            for col_num in range(3, len(row_values)):
                position = (col_num - 3) % 3
                if position == 1:
                    self.ws.cell(row=current_row, column=col_num).fill = self.commission_fill
                elif position == 2:
                    self.ws.cell(row=current_row, column=col_num).fill = self.net_after_fill
            self.ws.cell(row=current_row, column=len(row_values)).fill = self.net_after_fill

            current_row += 1
            grand_total_after_commission += total_after_commission

        total_row = ["Grand Total", ""]
        for period in periods:
            for platform in self.PLATFORM_GROUPS:
                platform_totals = period_totals_by_platform[period][platform["key"]]
                total_row.extend(
                    [
                        platform_totals["net"],
                        platform_totals["commission"],
                        platform_totals["after"],
                    ]
                )
        total_row.append(grand_total_after_commission)
        self.ws.append(total_row)

        total_row_index = self.ws.max_row
        for col_num in range(1, len(total_row) + 1):
            cell = self.ws.cell(row=total_row_index, column=col_num)
            cell.font = self.header_font
            if col_num <= 2:
                cell.alignment = self.center_align
            else:
                cell.style = "management_commission_currency_style"
            cell.fill = self.header_fill

        for col_num in range(3, len(total_row)):
            position = (col_num - 3) % 3
            if position == 1:
                self.ws.cell(row=total_row_index, column=col_num).fill = self.commission_fill
            elif position == 2:
                self.ws.cell(row=total_row_index, column=col_num).fill = self.net_after_fill
        self.ws.cell(row=total_row_index, column=len(total_row)).fill = self.net_after_fill

    def _set_column_widths(self):
        self.ws.column_dimensions[get_column_letter(1)].width = 35
        self.ws.column_dimensions[get_column_letter(2)].width = 15

        last_column = self.ws.max_column
        for col_num in range(3, last_column + 1):
            self.ws.column_dimensions[get_column_letter(col_num)].width = 18
