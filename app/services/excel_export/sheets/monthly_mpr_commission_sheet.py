import calendar

from openpyxl.styles import Alignment, Font, NamedStyle, PatternFill
from openpyxl.utils import get_column_letter

from app.services.excel_export.base_sheet import BaseSheet
from app.services.excel_export import mpr_calculations as mpr_calc


class MonthlyMprCommissionSheet(BaseSheet):
    PLATFORM_GROUPS = (
        {
            "key": "gojek",
            "label": "Gojek",
            "net_key": "gojek_net",
            "commission_key": "gojek_commission",
            "after_key": "gojek_net_after_commission",
            "rate": 1 - mpr_calc.MPR_STANDARD_NET_RATE,
        },
        {
            "key": "grab",
            "label": "Grab",
            "net_key": "grab_net",
            "commission_key": "grab_commission",
            "after_key": "grab_net_after_commission",
            "rate": 1 - mpr_calc.MPR_STANDARD_NET_RATE,
        },
        {
            "key": "shopee",
            "label": "Shopee",
            "net_key": "shopee_net",
            "commission_key": "shopee_commission",
            "after_key": "shopee_net_after_commission",
            "rate": 1 - mpr_calc.MPR_SHOPEE_NET_RATE,
        },
        {
            "key": "shopeepay",
            "label": "ShopeePay",
            "net_key": "shopeepay_net",
            "commission_key": "shopeepay_commission",
            "after_key": "shopeepay_net_after_commission",
            "rate": 1 - mpr_calc.MPR_QRIS_OVO_NET_RATE,
        },
        {
            "key": "tiktok",
            "label": "TikTok",
            "net_key": "tiktok_net",
            "commission_key": "tiktok_commission",
            "after_key": "tiktok_net_after_commission",
            "rate": 1 - mpr_calc.MPR_TIKTOK_NET_RATE,
        },
    )
    PERIOD_COLUMN_SPAN = len(PLATFORM_GROUPS) * 3

    def __init__(self, workbook, data):
        super().__init__(workbook, "Monthly MPR Commission", data)

    def generate(self):
        self._prepare_styles()
        self._create_header()
        self._populate_data()
        self._set_column_widths()
        self.ws.freeze_panes = "B3"

    def _prepare_styles(self):
        self.header_font = Font(bold=True)
        self.center_align = Alignment(horizontal="center", vertical="center")
        self.header_fill = PatternFill(
            start_color="D9EAD3",
            end_color="D9EAD3",
            fill_type="solid",
        )
        self.commission_fill = PatternFill(
            start_color="FFF2CC",
            end_color="FFF2CC",
            fill_type="solid",
        )
        self.net_after_fill = PatternFill(
            start_color="DDEBF7",
            end_color="DDEBF7",
            fill_type="solid",
        )
        self.currency_style = NamedStyle(
            name="mpr_currency_style",
            number_format="#,##0.00",
        )
        if "mpr_currency_style" not in self.wb.style_names:
            self.wb.add_named_style(self.currency_style)

    def _periods(self):
        if isinstance(self.data, dict) and "periods" in self.data:
            return self.data["periods"]
        return list(range(1, 13))

    def _outlets(self):
        if isinstance(self.data, dict) and "outlets" in self.data:
            return self.data.get("outlets", {})
        return self.data

    def _period_label(self, period):
        if isinstance(period, tuple):
            year, month = period
            return f"{calendar.month_name[month]} {year}"
        return calendar.month_name[period]

    def _platform_values(self, period_totals, platform):
        net_total = float(period_totals.get(platform["net_key"], 0) or 0)
        if not net_total and platform["net_key"] not in period_totals:
            return 0, 0, 0

        commission_total = float(
            period_totals.get(
                platform["commission_key"],
                net_total * platform["rate"],
            )
            or 0
        )
        net_after_commission = float(
            period_totals.get(
                platform["after_key"],
                net_total - commission_total,
            )
            or 0
        )
        return net_total, commission_total, net_after_commission

    def _create_header(self):
        periods = self._periods()

        self.ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        self.ws["A1"] = "Outlet"

        current_col = 2
        for period in periods:
            self.ws.merge_cells(
                start_row=1,
                start_column=current_col,
                end_row=1,
                end_column=current_col + self.PERIOD_COLUMN_SPAN - 1,
            )
            self.ws.cell(row=1, column=current_col, value=self._period_label(period))
            for platform_index, platform in enumerate(self.PLATFORM_GROUPS):
                platform_col = current_col + (platform_index * 3)
                self.ws.cell(row=2, column=platform_col, value=f"{platform['label']} Net")
                self.ws.cell(row=2, column=platform_col + 1, value="Commission")
                self.ws.cell(
                    row=2,
                    column=platform_col + 2,
                    value=f"{platform['label']} Net After Commission",
                )
            current_col += self.PERIOD_COLUMN_SPAN

        self.ws.merge_cells(
            start_row=1,
            start_column=current_col,
            end_row=2,
            end_column=current_col,
        )
        self.ws.cell(row=1, column=current_col, value="Total Commission")
        self.ws.merge_cells(
            start_row=1,
            start_column=current_col + 1,
            end_row=2,
            end_column=current_col + 1,
        )
        self.ws.cell(row=1, column=current_col + 1, value="Total After Commission")

        for row in self.ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=current_col + 1):
            for cell in row:
                cell.font = self.header_font
                cell.alignment = self.center_align
                cell.fill = self.header_fill

        for col_num in range(2, current_col):
            position = (col_num - 2) % 3
            if position == 1:
                self.ws.cell(row=2, column=col_num).fill = self.commission_fill
            elif position == 2:
                self.ws.cell(row=2, column=col_num).fill = self.net_after_fill
        self.ws.cell(row=1, column=current_col).fill = self.commission_fill
        self.ws.cell(row=1, column=current_col + 1).fill = self.net_after_fill

    def _populate_data(self):
        periods = self._periods()
        outlets = self._outlets()
        period_totals_by_platform = {
            period: {
                platform["key"]: {"net": 0, "commission": 0, "after": 0}
                for platform in self.PLATFORM_GROUPS
            }
            for period in periods
        }
        grand_total_commission = 0
        grand_total_after_commission = 0
        current_row = 3

        for _, outlet_data in outlets.items():
            row_values = [outlet_data["name"]]
            total_commission = outlet_data.get("total_commission", 0)
            total_after_commission = outlet_data.get("total_after_commission", 0)

            for period in periods:
                period_totals = outlet_data["monthly_totals"].get(
                    period,
                    {},
                )
                if not isinstance(period_totals, dict):
                    period_totals = {"net_total": period_totals}
                period_after_commission = 0
                for platform in self.PLATFORM_GROUPS:
                    net_total, commission_total, net_after_commission = self._platform_values(
                        period_totals,
                        platform,
                    )
                    period_totals_by_platform[period][platform["key"]]["net"] += net_total
                    period_totals_by_platform[period][platform["key"]]["commission"] += (
                        commission_total
                    )
                    period_totals_by_platform[period][platform["key"]]["after"] += (
                        net_after_commission
                    )
                    period_after_commission += net_after_commission
                    row_values.extend([net_total, commission_total, net_after_commission])

                if not total_after_commission:
                    total_after_commission += period_after_commission

            row_values.append(total_commission)
            row_values.append(total_after_commission)
            self.ws.append(row_values)

            for col_num in range(2, len(row_values) + 1):
                self.ws.cell(row=current_row, column=col_num).style = "mpr_currency_style"
            for col_num in range(2, len(row_values) - 1):
                position = (col_num - 2) % 3
                if position == 1:
                    self.ws.cell(row=current_row, column=col_num).fill = self.commission_fill
                elif position == 2:
                    self.ws.cell(row=current_row, column=col_num).fill = self.net_after_fill
            self.ws.cell(row=current_row, column=len(row_values) - 1).fill = self.commission_fill
            self.ws.cell(row=current_row, column=len(row_values)).fill = self.net_after_fill
            current_row += 1
            grand_total_commission += total_commission
            grand_total_after_commission += total_after_commission

        total_row = ["Grand Total"]
        for period in periods:
            for platform in self.PLATFORM_GROUPS:
                platform_totals = period_totals_by_platform[period][platform["key"]]
                total_row.extend(
                    [
                        platform_totals["net"],
                        platform_totals["commission"],
                        platform_totals["after"],
                    ]
                )
        total_row.append(grand_total_commission)
        total_row.append(grand_total_after_commission)
        self.ws.append(total_row)

        total_row_index = self.ws.max_row
        for col_num in range(1, len(total_row) + 1):
            cell = self.ws.cell(row=total_row_index, column=col_num)
            cell.font = self.header_font
            if col_num == 1:
                cell.alignment = self.center_align
            if col_num >= 2:
                cell.style = "mpr_currency_style"
            cell.fill = self.header_fill

        for col_num in range(2, len(total_row) - 1):
            position = (col_num - 2) % 3
            if position == 1:
                self.ws.cell(row=total_row_index, column=col_num).fill = self.commission_fill
            elif position == 2:
                self.ws.cell(row=total_row_index, column=col_num).fill = self.net_after_fill
        self.ws.cell(row=total_row_index, column=len(total_row) - 1).fill = self.commission_fill
        self.ws.cell(row=total_row_index, column=len(total_row)).fill = self.net_after_fill

    def _set_column_widths(self):
        self.ws.column_dimensions[get_column_letter(1)].width = 35

        last_column = self.ws.max_column
        for col_num in range(2, last_column + 1):
            self.ws.column_dimensions[get_column_letter(col_num)].width = 16
