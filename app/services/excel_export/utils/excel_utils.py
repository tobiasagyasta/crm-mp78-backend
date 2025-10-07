from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ======================================================================================
# FONT STYLES
# ======================================================================================
HEADER_FONT = Font(bold=True)
BOLD_RED_FONT = Font(bold=True, color='FF0000')

# ======================================================================================
# FILL STYLES
# ======================================================================================
YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
GRAY_FILL = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
LIGHT_RED_FILL = PatternFill(start_color='F4CCCC', end_color='F4CCCC', fill_type='solid')
LIGHT_GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
PALE_GREEN_FILL = PatternFill(start_color='98FB98', end_color='98FB98', fill_type='solid')
GOJEK_FILL = PatternFill(start_color='00AA13', end_color='00AA13', fill_type='solid')
GRAB_FILL = PatternFill(start_color='98FB98', end_color='98FB98', fill_type='solid')
SHOPEE_FILL = PatternFill(start_color='FF7A00', end_color='FF7A00', fill_type='solid')
TIKTOK_FILL = PatternFill(start_color='F227F5', end_color='F227F5', fill_type='solid')
BLUE_FILL = PatternFill(start_color='27A3F5', end_color='27A3F5', fill_type='solid')
TEAL_FILL = PatternFill(start_color='35F0F0', end_color='35F0F0', fill_type='solid')
CASH_FILL = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
COMMISSION_FILL = PatternFill(start_color='C6CCB2', end_color='C6CCB2', fill_type='solid')
SHOPEEPAY_FILL = PatternFill(start_color='E31F26', end_color='E31F26', fill_type='solid')
MUTATION_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
DIFFERENCE_FILL = PatternFill(start_color='F4CCCC', end_color='F4CCCC', fill_type='solid')
DATE_FILL = PatternFill(start_color='C9F0FF', end_color='C9F0FF', fill_type='solid')
SISA_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
DATA_FILL = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

# ======================================================================================
# ALIGNMENT STYLES
# ======================================================================================
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
RIGHT_ALIGN = Alignment(horizontal='right')

# ======================================================================================
# BORDER STYLES
# ======================================================================================
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ======================================================================================
# HELPER FUNCTIONS
# ======================================================================================
def set_column_widths(ws, widths: dict):
    """
    Sets the width for specified columns.
    :param ws: The worksheet object.
    :param widths: A dictionary mapping column letters to widths.
    """
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

def auto_fit_columns(ws):
    """
    Auto-fits all columns in a worksheet based on content length.
    """
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    # Add a little extra padding
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width